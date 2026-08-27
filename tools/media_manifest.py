#!/usr/bin/env python3
"""Asset Manifest v2 — deterministic media/production-asset compiler.

WHAT IS CANONICAL
-----------------
The CURRENT Markdown under ``02 Tervezet/`` is the single source of truth. Asset
requirements are declared inside the lesson/peula/hub files themselves, in hidden
HTML comments carrying strict JSON::

    <!-- @asset
    {"id": "M5.1-VID-01", "kind": "video", ...}
    -->

Copy that must be produced verbatim (narration, alt text, printable body text)
lives in the lesson file inside a *source block*::

    <!-- @source {"id": "M5.1-HOOK-VO", "kind": "narration"} -->
    > „Szia!
    > ...
    <!-- @endsource -->

Everything else — the v2 JSON manifest, the CSVs, the XLSX workbook, the Markdown
register, the migration map — is GENERATED from those declarations. Never edit a
generated file by hand; edit the Markdown and rebuild.

WHY
---
The predecessor system extracted asset rows with an AI workflow into a frozen
JSON snapshot. Every later curriculum edit silently drifted the stored narration,
alt text and line references away from the lessons, and nothing could detect it.
This compiler removes the snapshot: source text is read live from the Markdown on
every build, and ``check`` fails when a committed generated file no longer matches
what the current Markdown produces.

The canonical build is offline and deterministic. No network, no AI, no clock.

CLI
---
    python3 tools/media_manifest.py build       # regenerate every derived output
    python3 tools/media_manifest.py validate    # schema/reference/a11y checks only
    python3 tools/media_manifest.py check       # fail if a generated file is stale
    python3 tools/media_manifest.py reconcile   # legacy 747-row accounting
    python3 tools/media_manifest.py stats       # current counts
    python3 tools/media_manifest.py lint        # possible undeclared assets
    python3 tools/media_manifest.py diff <ref>  # what changed against a git ref
    python3 tools/media_manifest.py --selftest  # parser/normaliser smoke test
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import os
import re
import subprocess
import sys
import unicodedata
import zipfile
from pathlib import Path

SCHEMA_VERSION = "2.0"

ROOT = Path(__file__).resolve().parents[1]
ACTIVE_ROOT = ROOT / "02 Tervezet"
MODULE_ROOT = ACTIVE_ROOT / "Modulok"
MEDIA_ROOT = ACTIVE_ROOT / "Média-assetek"
BUILD_ROOT = MEDIA_ROOT / "_build"
LEGACY_ROOT = MEDIA_ROOT / "_legacy"

# --------------------------------------------------------------------------
# Discovery
#
# There is deliberately NO hard-coded list of lesson files. The old extractor
# carried a literal 74-entry inventory that pointed at a directory layout which
# no longer exists, which is exactly why it stopped being runnable. Discovery
# here is a directory walk plus a small, central exclusion list — a new lesson
# file is picked up with no Python edit.
# --------------------------------------------------------------------------

#: Directories under ``02 Tervezet`` that are NOT authoring sources. Only the
#: media register itself: it holds generated views plus its own documentation.
EXCLUDED_DIRS: tuple[str, ...] = ("02 Tervezet/Média-assetek",)

#: Files scanned by the discovery lint. The register carries assets for the
#: module corpus; the program-level policy documents (accessibility standard,
#: release gates, glossary) legitimately *talk about* captions and alt text
#: without owning any deliverable, so linting them would only produce noise.
LINT_ROOTS: tuple[str, ...] = ("02 Tervezet/Modulok",)

MODULE_ORDER = ("M0", "M1", "M2", "M3", "M4", "M5", "M6", "M7", "Z")
FILE_KIND_ORDER = ("hub", "kapu", "online-lecke", "peula", "program-doc")

# --------------------------------------------------------------------------
# Schema vocabularies
# --------------------------------------------------------------------------

KINDS = (
    "video", "voiceover", "audio", "animation",
    "image", "illustration", "diagram", "icon-set", "screenshot", "photo",
    "document", "template", "download", "print", "worksheet", "card-set", "poster",
    "other",
)

#: Subtypes are validated only for kinds where a closed set is genuinely useful.
SUBTYPES: dict[str, tuple[str, ...]] = {
    "video": ("ai-talking-head", "explainer", "interactive", "screen-recording", "live-action"),
    "voiceover": ("narration", "dialogue"),
    "audio": ("narration", "music", "sfx", "ambience"),
    "document": ("pdf", "docx", "spreadsheet", "checklist", "handout", "template"),
}

MODES = ("generate", "reuse", "external", "provided", "human-decision")
MODE_LABELS = {
    "generate": "legyártandó",
    "reuse": "újrahasznosítás",
    "external": "külső forrás",
    "provided": "meglévő szervezeti anyag",
    "human-decision": "emberi döntés kell",
}

STATUSES = (
    "spec-ready", "blocked", "pending-rights", "pending-human-decision",
    "pending-runtime", "pending-production-rule",
)
STATUS_LABELS = {
    "spec-ready": "specifikáció kész",
    "blocked": "blokkolt",
    "pending-rights": "jogtisztázás alatt",
    "pending-human-decision": "emberi döntésre vár",
    "pending-runtime": "runtime-döntésre vár",
    "pending-production-rule": "produkciós szabályra vár",
}

PROVENANCE = ("human", "ai", "stock", "third-party", "mixed", "unknown", "pending")
PROVENANCE_LABELS = {
    "human": "emberi",
    "ai": "AI-generált",
    "stock": "stock",
    "third-party": "harmadik fél",
    "mixed": "vegyes",
    "unknown": "ismeretlen",
    "pending": "eldöntetlen",
}

DERIVATIVES = (
    "voiceover", "captions", "transcript", "alt-text",
    "thumbnail", "audio-only", "low-bandwidth", "print-pdf", "editable",
)
DERIVATIVE_LABELS = {
    "voiceover": "felmondott hang",
    "captions": "felirat",
    "transcript": "leirat",
    "alt-text": "alt-szöveg",
    "thumbnail": "thumbnail",
    "audio-only": "csak hang változat",
    "low-bandwidth": "alacsony sávszélességű változat",
    "print-pdf": "nyomtatható PDF",
    "editable": "szerkeszthető, kitölthető változat",
}
#: Derivative role -> deliverable ID suffix after ``::``.
DERIVATIVE_SUFFIX = {d: d.upper().replace("-", "") for d in DERIVATIVES}

SOURCE_KINDS = ("narration", "alt-text", "caption", "transcript", "copy", "document")

#: Visual kinds must state whether they carry information or are decorative.
VISUAL_KINDS = ("image", "illustration", "diagram", "icon-set", "screenshot", "photo", "animation")
#: Kinds whose deliverable is audio-only and therefore needs a text equivalent.
AUDIO_KINDS = ("voiceover", "audio")
#: Print/document kinds.
PRINT_KINDS = ("print", "worksheet", "card-set", "poster")
DOC_KINDS = ("document", "template", "download")

#: Which kinds may reuse which. Reuse means the SAME produced file is used again,
#: so the kinds must be interchangeable in production.
REUSE_COMPATIBLE: dict[str, tuple[str, ...]] = {
    "image": ("image", "illustration", "photo", "screenshot", "diagram", "icon-set"),
    "illustration": ("illustration", "image", "diagram"),
    "diagram": ("diagram", "illustration", "image"),
    "icon-set": ("icon-set", "image", "illustration"),
    "photo": ("photo", "image"),
    "screenshot": ("screenshot", "image"),
    # A poster and a worksheet are not interchangeable produced files: a wall
    # poster reuse of a fill-in worksheet hid a safeguarding step-map mismatch.
    "print": PRINT_KINDS,
    "worksheet": ("worksheet", "card-set", "print", "document", "template", "download"),
    "card-set": ("card-set", "worksheet", "print"),
    "poster": ("poster", "print"),
    "document": DOC_KINDS + ("worksheet",),
    "template": DOC_KINDS + ("worksheet",),
    "download": DOC_KINDS + ("worksheet",),
    "video": ("video", "animation"),
    "animation": ("animation", "video"),
    "voiceover": ("voiceover", "audio"),
    "audio": ("audio", "voiceover"),
    "other": ("other",),
}

REQUIRED_ASSET_FIELDS = ("id", "kind", "title")

KNOWN_ASSET_FIELDS = {
    "id", "unit", "kind", "subtype", "mode", "title", "purpose", "spec",
    "source_ref", "composed_of", "provenance", "provenance_note", "technical",
    "a11y", "derivatives", "reuse_of", "external", "status", "blockers",
    "production_rules", "decision", "notes", "legacy", "review",
}
KNOWN_SOURCE_FIELDS = {"id", "kind", "note", "for"}
KNOWN_ASSET_FREE_FIELDS = {"reason"}
KNOWN_A11Y_FIELDS = {"visual", "audio", "alt_source_ref", "alt_note", "note"}
KNOWN_EXTERNAL_FIELDS = {"source", "url", "path", "owner", "licence", "evidence", "replace"}

# --------------------------------------------------------------------------
# Generated output paths
# --------------------------------------------------------------------------

OUT_JSON = BUILD_ROOT / "media-manifest.v2.json"
OUT_ASSETS_CSV = MEDIA_ROOT / "assetek.csv"
OUT_DELIVERABLES_CSV = MEDIA_ROOT / "deliverable-ek.csv"
OUT_REUSE_CSV = MEDIA_ROOT / "ujrahasznositas.csv"
OUT_REGISTER_MD = MEDIA_ROOT / "Média-asset regiszter.md"
OUT_XLSX = MEDIA_ROOT / "Média-asset regiszter.xlsx"
OUT_MIGRATION_CSV = MEDIA_ROOT / "asset-migration-map.csv"
OUT_MIGRATION_MD = MEDIA_ROOT / "ASSET-MANIFEST-V2-MIGRATION.md"

LEGACY_JSON = LEGACY_ROOT / "media-merged.json"
LEGACY_JSON_FALLBACK = BUILD_ROOT / "media-merged.json"
LEGACY_DISPOSITIONS = LEGACY_ROOT / "legacy-dispositions.json"

#: Every output the build writes, in build order. ``check`` compares all of them.
GENERATED_OUTPUTS = (
    OUT_JSON, OUT_ASSETS_CSV, OUT_DELIVERABLES_CSV, OUT_REUSE_CSV,
    OUT_REGISTER_MD, OUT_XLSX, OUT_MIGRATION_CSV, OUT_MIGRATION_MD,
)


# ==========================================================================
# Errors
# ==========================================================================

class ManifestError(Exception):
    """A structured, actionable manifest problem."""

    def __init__(self, where: str, message: str, fix: str = ""):
        self.where = where
        self.message = message
        self.fix = fix
        super().__init__(str(self))

    def __str__(self) -> str:
        tail = f"  → {self.fix}" if self.fix else ""
        return f"{self.where}: {self.message}{tail}"


# ==========================================================================
# Text normalisation and hashing
# ==========================================================================

def normalise_source_text(text: str) -> str:
    """Canonical form of a source block body.

    Deterministic and documented, because the copy hash is built from it:

    * Unicode NFC (the corpus mixes precomposed and decomposed Hungarian
      diacritics after its AFFiNE export);
    * CRLF/CR collapsed to LF;
    * the Markdown blockquote marker ``>`` (and one following space) stripped
      from the start of each line — the compiler removes *markup*, never words;
    * trailing whitespace removed per line;
    * leading/trailing blank lines removed.

    Nothing else. No paraphrase, no re-wrapping, no punctuation "cleanup":
    quotation marks, emphasis markers and paragraph breaks all survive.
    """
    text = unicodedata.normalize("NFC", text)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    out = []
    for line in text.split("\n"):
        stripped = line.lstrip()
        if stripped.startswith(">"):
            stripped = stripped[1:]
            if stripped.startswith(" "):
                stripped = stripped[1:]
            line = stripped
        out.append(line.rstrip())
    while out and not out[0].strip():
        out.pop(0)
    while out and not out[-1].strip():
        out.pop()
    return "\n".join(out)


#: A source block is a verbatim region of the lesson. Where an asset needs only
#: the quotation inside it — a prescribed alt written as `**Javasolt alt:** „…”`
#: within a longer accessibility instruction — the REFERENCE carries the
#: selector: `"M1.1-DIA-01-ALT#1"` takes that quotation. Selection only; the
#: compiler never rewrites the words.
#:
#: The selector is deliberately NOT positional. An alt source addressed with `#n`
#: must hold exactly ONE „…” quotation and the selector must be `#1`; a second
#: quotation is a validation error rather than a silent choice. Positional
#: selection had a failure mode with no symptom: inserting an unrelated quotation
#: earlier in the block rebinds the alt to the wrong text while CI stays green.
#: Where the alt cannot be wrapped that tightly without splitting a rendered
#: paragraph, the asset carries `a11y.alt_note` instead and is not live-bound.
QUOTED_SPAN = re.compile(r"„(?P<text>[^„”]*)”", re.S)


def split_ref(ref: str) -> tuple[str, int | None]:
    if "#" in ref:
        source_id, _, index = ref.rpartition("#")
        if source_id and index.isdigit():
            return source_id, int(index)
    return ref, None


def select_quote(text: str, index: int, where: str, ref: str) -> str:
    matches = QUOTED_SPAN.findall(text)
    if not matches:
        raise ManifestError(where, f"a(z) {ref} idézetre hivatkozik, de a forrásban "
                                   "nincs „…” idézet",
                            "vedd le a `#n` szelektort, vagy jelöld idézőjellel a szöveget")
    if not 1 <= index <= len(matches):
        raise ManifestError(where, f"a(z) {ref} a(z) {index}. idézetet kéri, "
                                   f"de a forrásban {len(matches)} van",
                            f"használj 1 és {len(matches)} közötti sorszámot")
    return matches[index - 1].strip()


def resolve_ref(sources: dict, ref: str, where: str) -> tuple[dict | None, str]:
    """(source record, selected text) for a possibly `#n`-qualified reference."""
    source_id, index = split_ref(ref)
    src = sources.get(source_id)
    if src is None:
        return None, ""
    if index is None:
        return src, src["text"]
    return src, select_quote(src["text"], index, where, ref)


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def canonical_json(value) -> str:
    """Stable JSON for hashing: sorted keys, no incidental whitespace."""
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def natural_key(value: str):
    return [int(part) if part.isdigit() else part.casefold()
            for part in re.split(r"(\d+)", value or "")]


# ==========================================================================
# File discovery and unit identity
# ==========================================================================

#: Repository root used for relative paths. ``diff`` retargets it at a temporary
#: checkout of an older ref; everything else leaves it at the real repository.
_REL_ROOT = ROOT


class _rel_root:
    """Temporarily resolve relative paths against another root (used by ``diff``)."""

    def __init__(self, root: Path):
        self.root = root

    def __enter__(self):
        global _REL_ROOT
        self.previous = _REL_ROOT
        _REL_ROOT = self.root
        return self

    def __exit__(self, *exc):
        global _REL_ROOT
        _REL_ROOT = self.previous
        return False


def _rel(path: Path) -> str:
    """Repository-relative POSIX path, NFC-normalised.

    macOS hands out decomposed (NFD) filenames, so an archive unpacked there
    would not match the NFC strings the exclusion list and the unit rules are
    written in — the build would fail with a cascade of unrelated errors.
    """
    try:
        rel = path.relative_to(_REL_ROOT).as_posix()
    except ValueError:
        rel = path.as_posix()
    return unicodedata.normalize("NFC", rel)


def discover_sources(root: Path = ACTIVE_ROOT) -> list[Path]:
    """Every current authoring Markdown file, deterministically ordered."""
    files = []
    for path in root.rglob("*.md"):
        rel = _rel(path)
        if any(rel == d or rel.startswith(d + "/") for d in EXCLUDED_DIRS):
            continue
        files.append(path)
    return sorted(files, key=lambda p: _rel(p))


def in_lint_scope(path: Path) -> bool:
    rel = _rel(path)
    return any(rel == d or rel.startswith(d + "/") for d in LINT_ROOTS)


_UNIT_CODE = re.compile(r"^((?:M[0-7]|Z)(?:\.[0-9A-ZÁÉÍÓÖŐÚÜŰ]+)?)\b")


def file_identity(path: Path) -> tuple[str, str, str]:
    """(unit, module, file kind) for one source file — derived from the path.

    The v2 unit token also resolves the historical ID-space collision: a module
    hub file that repeats a lesson's asset gets its own ``M1-HUB`` namespace
    instead of re-using the lesson's ``M1.1`` identifier for a different row.
    """
    rel = _rel(path)
    stem = path.stem
    parts = tuple(rel.split("/"))

    if len(parts) >= 3 and parts[1] == "Modulok":
        module = parts[2]
        if len(parts) >= 5 and parts[3] in ("Online leckék", "Peulák"):
            match = _UNIT_CODE.match(stem)
            unit = match.group(1) if match else stem.split(" ")[0]
            kind = "online-lecke" if parts[3] == "Online leckék" else "peula"
            return unit, module, kind
        if "Kapu" in stem:
            return f"{module}-KAPU", module, "kapu"
        return f"{module}-HUB", module, "hub"

    slug = re.split(r"\s+[–—-]\s+", stem)[0]
    slug = unicodedata.normalize("NFKD", slug)
    slug = "".join(ch for ch in slug if not unicodedata.combining(ch))
    slug = re.sub(r"[^A-Za-z0-9]+", "", slug).upper() or "DOC"
    if rel.startswith("02 Tervezet/") and rel.count("/") == 1:
        return slug, "—", "program-doc"
    return slug, "—", "program-doc"


def module_sort_index(module: str) -> int:
    try:
        return MODULE_ORDER.index(module)
    except ValueError:
        return len(MODULE_ORDER)


def file_kind_index(kind: str) -> int:
    try:
        return FILE_KIND_ORDER.index(kind)
    except ValueError:
        return len(FILE_KIND_ORDER)


# ==========================================================================
# Declaration parser
# ==========================================================================

#: A marker may carry a prefix so it can sit INSIDE the block it wraps: spaces
#: to stay within a list item, `>` to stay within a blockquote. Without that, a
#: marker between two list items ends the list and a marker between two quoted
#: lines splits the quote box — both visible to the reader.
_OPEN_RE = re.compile(r"^(?P<prefix>[ \t>]*)<!--\s*@(?P<tag>[a-z][a-z0-9_-]*)\b(?P<rest>.*)$")
_ENDSOURCE_RE = re.compile(r"^[ \t>]*<!--\s*@endsource\s*-->\s*$")
#: A declaration payload may not run longer than this many lines. Guards against
#: an unclosed comment swallowing the rest of a lesson.
MAX_DECL_LINES = 400


def _parse_payload(lines: list[str], start: int, tag: str, where: str):
    """Read one ``<!-- @tag {json} -->`` declaration starting at ``lines[start]``.

    Returns ``(payload_dict, end_index)`` where ``end_index`` is the index of the
    closing line. The closing ``-->`` may sit on the opening line or on any later
    line; the first candidate whose accumulated payload parses as JSON wins, so a
    literal ``-->`` inside a JSON string does not truncate the block.
    """
    first = _OPEN_RE.match(lines[start])
    assert first is not None
    rest = first.group("rest")
    chunks: list[str] = []
    last_error = None

    def try_close(body_parts: list[str], tail: str):
        candidate = "\n".join(body_parts + [tail]).strip()
        if not candidate:
            return {}
        return json.loads(candidate)

    # Closing on the opening line.
    stripped = rest.rstrip()
    if stripped.endswith("-->"):
        try:
            payload = try_close([], stripped[: -len("-->")])
            return payload, start
        except json.JSONDecodeError as exc:
            last_error = exc
    chunks.append(rest)

    for idx in range(start + 1, min(len(lines), start + MAX_DECL_LINES)):
        line = lines[idx]
        stripped = line.rstrip()
        if stripped.endswith("-->"):
            try:
                payload = try_close(chunks, stripped[: -len("-->")])
                return payload, idx
            except json.JSONDecodeError as exc:
                last_error = exc
        chunks.append(line)

    if last_error is not None:
        raise ManifestError(
            f"{where}:{start + 1}",
            f"a @{tag} deklaráció JSON-je hibás: {last_error}",
            "javítsd a JSON-t (szigorú JSON: dupla idézőjel, nincs záró vessző)",
        )
    raise ManifestError(
        f"{where}:{start + 1}",
        f"lezáratlan @{tag} deklaráció (nincs `-->` {MAX_DECL_LINES} soron belül)",
        "zárd le a blokkot egy `-->` sorral",
    )


def parse_file(path: Path) -> dict:
    """Parse one Markdown file into raw declarations and source blocks."""
    where = _rel(path)
    text = path.read_text(encoding="utf-8")
    lines = text.split("\n")

    assets: list[dict] = []
    sources: list[dict] = []
    asset_free: list[dict] = []
    #: line ranges (1-based, inclusive) occupied by v2 metadata, for the lint
    covered: list[tuple[int, int]] = []

    idx = 0
    while idx < len(lines):
        line = lines[idx]
        match = _OPEN_RE.match(line)
        if not match:
            idx += 1
            continue
        tag = match.group("tag")

        if tag == "endsource":
            if not _ENDSOURCE_RE.match(line):
                raise ManifestError(
                    f"{where}:{idx + 1}",
                    "hibás @endsource sor",
                    "pontosan ennyi legyen a sor: <!-- @endsource -->",
                )
            raise ManifestError(
                f"{where}:{idx + 1}",
                "@endsource nyitó @source blokk nélkül",
                "vagy törölt @source fejlécet pótolj, vagy töröld ezt a sort",
            )

        if tag not in ("asset", "source", "asset-free"):
            idx += 1
            continue

        payload, end = _parse_payload(lines, idx, tag, where)
        if not isinstance(payload, dict):
            raise ManifestError(
                f"{where}:{idx + 1}",
                f"a @{tag} deklaráció nem JSON-objektum",
                "kapcsos zárójeles objektumot adj meg",
            )

        if tag == "asset-free":
            payload["_file"] = where
            payload["_line"] = idx + 1
            asset_free.append(payload)
            covered.append((idx + 1, end + 1))
            idx = end + 1
            continue

        if tag == "asset":
            payload["_file"] = where
            payload["_line"] = idx + 1
            assets.append(payload)
            covered.append((idx + 1, end + 1))
            idx = end + 1
            continue

        # @source: body runs until the matching @endsource.
        body_start = end + 1
        body_end = None
        for scan in range(body_start, len(lines)):
            if _ENDSOURCE_RE.match(lines[scan]):
                body_end = scan
                break
        if body_end is None:
            raise ManifestError(
                f"{where}:{idx + 1}",
                f"lezáratlan @source blokk ({payload.get('id', '?')})",
                "tedd ki a záró <!-- @endsource --> sort",
            )
        payload["_file"] = where
        payload["_line"] = idx + 1
        payload["_body"] = "\n".join(lines[body_start:body_end])
        payload["_body_lines"] = (body_start + 1, body_end)
        sources.append(payload)
        covered.append((idx + 1, body_end + 1))
        idx = body_end + 1

    return {"path": path, "rel": where, "lines": lines, "assets": assets,
            "sources": sources, "asset_free": asset_free, "covered": covered}


# ==========================================================================
# Model construction
# ==========================================================================

def _as_list(value, field: str, where: str) -> list:
    if value is None:
        return []
    if not isinstance(value, list):
        raise ManifestError(where, f"a `{field}` mező listát vár", "használj JSON tömböt")
    return value


def _check_unknown(payload: dict, allowed: set[str], where: str, what: str) -> None:
    unknown = sorted(k for k in payload if not k.startswith("_") and k not in allowed)
    if unknown:
        raise ManifestError(
            where,
            f"ismeretlen {what} mező: {', '.join(unknown)}",
            f"engedélyezett mezők: {', '.join(sorted(allowed))}",
        )


def _validate_asset_free(parsed_files: list[dict], errors: list[ManifestError]) -> None:
    """`@asset-free` states, with a reason, that a file owns no deliverable.

    It exists so "this file has nothing to produce" is an explicit, reviewable
    claim rather than the silence the discovery lint cannot tell from an omission.
    """
    for parsed in parsed_files:
        declarations = parsed["asset_free"]
        if not declarations:
            continue
        for raw in declarations:
            where = f"{raw['_file']}:{raw['_line']}"
            try:
                _check_unknown(raw, KNOWN_ASSET_FREE_FIELDS, where, "@asset-free")
                if not raw.get("reason"):
                    raise ManifestError(where, "az @asset-free deklarációhoz kötelező a `reason`",
                                        '{"reason": "<miért nincs itt legyártandó anyag>"}')
            except ManifestError as exc:
                errors.append(exc)
        if len(declarations) > 1:
            errors.append(ManifestError(
                f"{declarations[1]['_file']}:{declarations[1]['_line']}",
                "egy fájlban több @asset-free deklaráció van", "tarts meg egyet"))
        if parsed["assets"]:
            errors.append(ManifestError(
                f"{declarations[0]['_file']}:{declarations[0]['_line']}",
                "a fájl @asset-free, mégis van benne @asset deklaráció",
                "vagy töröld az @asset-free sort, vagy vidd át az assetet a helyére"))


def build_sources(parsed_files: list[dict], errors: list[ManifestError]) -> dict[str, dict]:
    sources: dict[str, dict] = {}
    for parsed in parsed_files:
        unit, module, file_kind = file_identity(parsed["path"])
        for raw in parsed["sources"]:
            where = f"{raw['_file']}:{raw['_line']}"
            try:
                _check_unknown(raw, KNOWN_SOURCE_FIELDS, where, "@source")
                sid = raw.get("id")
                if not isinstance(sid, str) or not sid.strip():
                    raise ManifestError(where, "a @source blokknak kötelező az `id`-je",
                                        'pl. {"id": "M5.1-HOOK-VO", "kind": "narration"}')
                if "::" in sid:
                    raise ManifestError(where, f"a forrás-ID nem tartalmazhat `::`-t: {sid}",
                                        "a `::` a deliverable-ID-k fenntartott elválasztója")
                kind = raw.get("kind")
                if kind not in SOURCE_KINDS:
                    raise ManifestError(where, f"ismeretlen forrás-típus: {kind!r}",
                                        f"engedélyezett: {', '.join(SOURCE_KINDS)}")
                if sid in sources:
                    prev = sources[sid]
                    raise ManifestError(where, f"duplikált forrás-ID: {sid}",
                                        f"már használatban itt: {prev['file']}:{prev['line']}")
                body = normalise_source_text(raw["_body"])
                if not body.strip():
                    raise ManifestError(where, f"üres @source blokk: {sid}",
                                        "vagy tedd bele a szöveget, vagy töröld a blokkot")
                sources[sid] = {
                    "id": sid,
                    "kind": kind,
                    "file": raw["_file"],
                    "line": raw["_line"],
                    "body_lines": list(raw["_body_lines"]),
                    "unit": unit,
                    "module": module,
                    "file_kind": file_kind,
                    "text": body,
                    "hash": sha256_text(body),
                    "note": raw.get("note", ""),
                    "for": raw.get("for", ""),
                    "used_by": [],
                }
            except ManifestError as exc:
                errors.append(exc)
    return sources


def build_assets(parsed_files: list[dict], errors: list[ManifestError]) -> dict[str, dict]:
    assets: dict[str, dict] = {}
    for parsed in parsed_files:
        unit, module, file_kind = file_identity(parsed["path"])
        for raw in parsed["assets"]:
            where = f"{raw['_file']}:{raw['_line']}"
            try:
                asset = _normalise_asset(raw, where, unit, module, file_kind)
            except ManifestError as exc:
                errors.append(exc)
                continue
            if asset["id"] in assets:
                prev = assets[asset["id"]]
                errors.append(ManifestError(
                    where, f"duplikált asset-ID: {asset['id']}",
                    f"már használatban itt: {prev['file']}:{prev['line']}"))
                continue
            assets[asset["id"]] = asset
    return assets


def _normalise_asset(raw: dict, where: str, unit: str, module: str, file_kind: str) -> dict:
    _check_unknown(raw, KNOWN_ASSET_FIELDS, where, "@asset")

    for field in REQUIRED_ASSET_FIELDS:
        if not raw.get(field):
            raise ManifestError(where, f"hiányzó kötelező mező: `{field}`",
                                f"kötelező mezők: {', '.join(REQUIRED_ASSET_FIELDS)}")

    aid = raw["id"]
    if not isinstance(aid, str) or "::" in aid:
        raise ManifestError(where, f"érvénytelen asset-ID: {aid!r}",
                            "a `::` a deliverable-ID-k fenntartott elválasztója")
    if not aid.startswith(unit + "-"):
        raise ManifestError(
            where, f"az asset-ID nem a fájl egységéhez tartozik: {aid} (várt előtag: {unit}-)",
            "az ID-nek a fájlból származtatott egység-előtaggal kell kezdődnie "
            "(lecke: M1.1-, peula: M1.A-, hub: M1-HUB-, kapu: M1-KAPU-)")
    declared_unit = raw.get("unit")
    if declared_unit is not None and declared_unit != unit:
        raise ManifestError(where, f"a deklarált `unit` ({declared_unit}) nem egyezik "
                                   f"a fájlból származtatottal ({unit})",
                            "hagyd el a `unit` mezőt — a fájl útvonala határozza meg")

    kind = raw["kind"]
    if kind not in KINDS:
        raise ManifestError(where, f"ismeretlen `kind`: {kind!r}",
                            f"engedélyezett: {', '.join(KINDS)}")
    subtype = raw.get("subtype")
    if subtype is not None:
        if kind in SUBTYPES and subtype not in SUBTYPES[kind]:
            raise ManifestError(where, f"ismeretlen `subtype` a(z) {kind} típushoz: {subtype!r}",
                                f"engedélyezett: {', '.join(SUBTYPES[kind])}")
        if not isinstance(subtype, str):
            raise ManifestError(where, "a `subtype` szöveg legyen", "")

    mode = raw.get("mode", "generate")
    if mode not in MODES:
        raise ManifestError(where, f"ismeretlen `mode`: {mode!r}",
                            f"engedélyezett: {', '.join(MODES)}")

    provenance = raw.get("provenance", "unknown")
    if provenance not in PROVENANCE:
        raise ManifestError(where, f"ismeretlen `provenance`: {provenance!r}",
                            f"engedélyezett: {', '.join(PROVENANCE)}")

    derivatives = _as_list(raw.get("derivatives"), "derivatives", where)
    for der in derivatives:
        if der not in DERIVATIVES:
            raise ManifestError(where, f"ismeretlen derivatíva: {der!r}",
                                f"engedélyezett: {', '.join(DERIVATIVES)}")
    if len(set(derivatives)) != len(derivatives):
        raise ManifestError(where, "ismétlődő derivatíva a listában", "minden szerep egyszer szerepeljen")

    a11y = raw.get("a11y") or {}
    if not isinstance(a11y, dict):
        raise ManifestError(where, "az `a11y` objektum legyen", '{"visual": "informative"}')
    _check_unknown(a11y, KNOWN_A11Y_FIELDS, where, "a11y")
    if a11y.get("visual") not in (None, "informative", "decorative"):
        raise ManifestError(where, f"ismeretlen a11y.visual: {a11y.get('visual')!r}",
                            "informative vagy decorative")
    if a11y.get("audio") not in (None, "spoken", "silent"):
        raise ManifestError(where, f"ismeretlen a11y.audio: {a11y.get('audio')!r}",
                            "spoken vagy silent")

    external = raw.get("external") or {}
    if not isinstance(external, dict):
        raise ManifestError(where, "az `external` objektum legyen", "")
    _check_unknown(external, KNOWN_EXTERNAL_FIELDS, where, "external")

    technical = raw.get("technical") or {}
    if not isinstance(technical, dict):
        raise ManifestError(where, "a `technical` objektum legyen", '{"aspect_ratio": "16:9"}')

    status = raw.get("status")
    if status is not None and status not in STATUSES:
        raise ManifestError(where, f"ismeretlen `status`: {status!r}",
                            f"engedélyezett: {', '.join(STATUSES)}")

    legacy = raw.get("legacy") or {}
    if not isinstance(legacy, dict):
        raise ManifestError(where, "a `legacy` objektum legyen (szerep → régi ID-k)", "")
    for role, ids in legacy.items():
        if role != "asset" and role not in DERIVATIVES:
            raise ManifestError(where, f"ismeretlen legacy-szerep: {role!r}",
                                f"engedélyezett: asset, {', '.join(DERIVATIVES)}")
        if not isinstance(ids, list) or not all(isinstance(i, str) for i in ids):
            raise ManifestError(where, f"a legacy.{role} szöveg-lista legyen", "")

    asset = {
        "id": aid,
        "unit": unit,
        "module": module,
        "file_kind": file_kind,
        "file": raw["_file"],
        "line": raw["_line"],
        "kind": kind,
        "subtype": subtype or "",
        "mode": mode,
        "title": raw["title"],
        "purpose": raw.get("purpose", ""),
        "spec": raw.get("spec", ""),
        "source_ref": raw.get("source_ref", ""),
        "composed_of": [str(c) for c in _as_list(raw.get("composed_of"),
                                                 "composed_of", where)],
        # Set during resolution: True only once every component's script actually
        # resolved. A declared composition is a promise, not a script.
        "composed_of_resolved": False,
        "provenance": provenance,
        "provenance_note": raw.get("provenance_note", ""),
        "technical": technical,
        "a11y": a11y,
        "derivatives": derivatives,
        "reuse_of": raw.get("reuse_of", ""),
        "external": external,
        "status": status,
        "blockers": [str(b) for b in _as_list(raw.get("blockers"), "blockers", where)],
        "production_rules": [str(r) for r in _as_list(raw.get("production_rules"),
                                                      "production_rules", where)],
        "decision": raw.get("decision", ""),
        "notes": raw.get("notes", ""),
        "review": raw.get("review", ""),
        "legacy": {role: list(ids) for role, ids in legacy.items()},
    }
    return asset


# ==========================================================================
# Validation
# ==========================================================================

def validate(assets: dict[str, dict], sources: dict[str, dict],
             errors: list[ManifestError]) -> None:
    _validate_references(assets, sources, errors)
    _validate_reuse(assets, errors)
    _validate_composition(assets, errors)
    _validate_accessibility(assets, errors)
    _validate_modes(assets, errors)
    _validate_sources_used(sources, errors)


def _validate_references(assets, sources, errors) -> None:
    for asset in assets.values():
        where = f"{asset['file']}:{asset['line']}"
        for cid in asset["composed_of"]:
            component = assets.get(cid)
            ref = component["source_ref"] if component else ""
            if ref and split_ref(ref)[0] in sources:
                sources[split_ref(ref)[0]]["used_by"].append(
                    f"{asset['id']}:composed_of")
        for field, ref in (("source_ref", asset["source_ref"]),
                           ("a11y.alt_source_ref", asset["a11y"].get("alt_source_ref", ""))):
            if not ref:
                continue
            src, _text = resolve_ref(sources, ref, where)
            if src is None:
                errors.append(ManifestError(
                    where, f"a(z) {asset['id']} `{field}` nem létező forrásra mutat: {ref}",
                    "hozd létre a @source blokkot, vagy javítsd a hivatkozást"))
                continue
            if src["file"] != asset["file"]:
                errors.append(ManifestError(
                    where, f"a(z) {asset['id']} `{field}` másik fájl forrására mutat "
                           f"({ref} itt: {src['file']})",
                    "a forrásblokk annak a fájlnak a szövegét írja le, ahol az asset él — "
                    "másold át a forrást, vagy használj explicit reuse_of-ot"))
                continue
            sources[split_ref(ref)[0]]["used_by"].append(f"{asset['id']}:{field}")
            if field == "source_ref" and asset["kind"] in ("video", "voiceover", "audio"):
                if src["kind"] not in ("narration", "caption", "transcript", "copy"):
                    errors.append(ManifestError(
                        where, f"a(z) {asset['id']} beszélt assethez nem beszéd-forrás "
                               f"tartozik ({ref}: {src['kind']})",
                        'a hang/videó forrása "narration" típusú legyen'))
            if field == "a11y.alt_source_ref":
                _, index = split_ref(ref)
                quotes = QUOTED_SPAN.findall(src["text"])
                if index is not None:
                    if len(quotes) != 1:
                        errors.append(ManifestError(
                            where, f"a(z) {asset['id']} alt-forrása {len(quotes)} „…” "
                                   f"idézetet tartalmaz ({ref}) — a pozíciós "
                                   "kiválasztás így elcsúszhat",
                            "szűkítsd a @source blokkot pontosan egy idézetre; "
                            "a szelektor csak #1 lehet"))
                    elif index != 1:
                        errors.append(ManifestError(
                            where, f"a(z) {asset['id']} alt-hivatkozása pozíciós "
                                   f"szelektort használ ({ref})",
                            "egy idézetes forrásnál a szelektor #1"))
            if field == "a11y.alt_source_ref" and src["kind"] != "alt-text":
                errors.append(ManifestError(
                    where, f"a(z) {asset['id']} alt-forrása nem alt-text típusú "
                           f"({ref}: {src['kind']})",
                    'az alt-szöveg forrásblokk kind-je "alt-text" legyen'))


def _validate_reuse(assets, errors) -> None:
    for asset in assets.values():
        where = f"{asset['file']}:{asset['line']}"
        target_id = asset["reuse_of"]
        if asset["mode"] == "reuse":
            if not target_id:
                errors.append(ManifestError(
                    where, f"a(z) {asset['id']} mode=reuse, de nincs `reuse_of`",
                    "add meg, melyik legyártott assetet használja újra"))
                continue
            if asset["derivatives"]:
                errors.append(ManifestError(
                    where, f"a(z) {asset['id']} újrahasznosítás, mégis vannak derivatívái",
                    "az újrahasznosított asset a kanonikus asset deliverable-jeit használja — "
                    "töröld a `derivatives` mezőt"))
        elif target_id:
            errors.append(ManifestError(
                where, f"a(z) {asset['id']} `reuse_of`-ot ad meg, de a mode nem reuse",
                'állítsd be: "mode": "reuse"'))
            continue
        if not target_id:
            continue
        if target_id == asset["id"]:
            errors.append(ManifestError(where, f"a(z) {asset['id']} önmagát hasznosítja újra",
                                        "a `reuse_of` másik assetre mutasson"))
            continue
        target = assets.get(target_id)
        if target is None:
            errors.append(ManifestError(
                where, f"a(z) {asset['id']} `reuse_of` nem létező assetre mutat: {target_id}",
                "javítsd az azonosítót, vagy hozd létre a kanonikus assetet"))
            continue
        compatible = REUSE_COMPATIBLE.get(asset["kind"], (asset["kind"],))
        if target["kind"] not in compatible:
            errors.append(ManifestError(
                where, f"a(z) {asset['id']} ({asset['kind']}) nem hasznosíthatja újra "
                       f"a(z) {target_id}-t ({target['kind']})",
                f"kompatibilis típusok: {', '.join(compatible)}"))

    # Cycles (and chains that never reach a produced asset).
    for asset in assets.values():
        if asset["mode"] != "reuse" or not asset["reuse_of"]:
            continue
        seen = [asset["id"]]
        current = assets.get(asset["reuse_of"])
        while current is not None and current["mode"] == "reuse" and current["reuse_of"]:
            if current["id"] in seen:
                errors.append(ManifestError(
                    f"{asset['file']}:{asset['line']}",
                    f"körkörös újrahasznosítás: {' → '.join(seen + [current['id']])}",
                    "az újrahasznosítási láncnak legyártott assetnél kell végződnie"))
                break
            seen.append(current["id"])
            current = assets.get(current["reuse_of"])


#: Wording that unambiguously demands a text equivalent. A visual may not be
#: declared decorative while its own note says an alt is required — that is the
#: one direction of this mistake that silently removes an access requirement.
_DEMANDS_ALT = re.compile(
    r"alt-?\s?sz[öo]veg (kell|jár|kötelező|szükséges)|érdemi alt|"
    r"nem dekorat|dekorat\w* NEM")


def composed_component_ids(assets) -> set[str]:
    """Every asset id that is delivered only inside a composed container."""
    values = assets.values() if isinstance(assets, dict) else assets
    return {cid for asset in values for cid in asset.get("composed_of", [])}


def _validate_composition(assets, errors) -> None:
    """`composed_of` — one runtime artefact assembled from several components.

    Grounded in the H5P Interactive Video content type itself: its `semantics.json`
    takes ONE base video (`files` / "Add a video", where the extra entries are
    alternative encodings of the same video) and ONE `textTracks` list of WebVTT
    tracks. Three scene clips therefore become one base video with exactly one
    caption track and one transcript — the components have nowhere to attach
    their own. So the container owns the text tracks and the components must not
    also declare them, and the container's script is the ordered concatenation of
    the components' own source blocks rather than a fourth copy of the same text.
    """
    for asset in assets.values():
        where = f"{asset['file']}:{asset['line']}"
        components = asset["composed_of"]
        if not components:
            continue
        if asset["kind"] != "video" or asset["a11y"].get("audio") != "spoken":
            errors.append(ManifestError(
                where, f"a(z) {asset['id']} `composed_of`-ot deklarál, de nem beszélt videó",
                "a kompozíció beszélt videó-konténerre való "
                '("kind": "video", "a11y": {"audio": "spoken"})'))
        if asset["mode"] != "generate":
            errors.append(ManifestError(
                where, f"a(z) {asset['id']} `composed_of`-ot deklarál, de a módja "
                       f"{asset['mode']}",
                'a kompozíció "generate" módú konténerre való'))
        if asset["source_ref"]:
            errors.append(ManifestError(
                where, f"a(z) {asset['id']} egyszerre deklarál `source_ref`-et és "
                       "`composed_of`-ot",
                "a konténer szkriptje vagy saját forrásblokk, vagy az összetevőké — "
                "ne legyen két igazsága"))
        for role in ("captions", "transcript"):
            if role not in asset["derivatives"]:
                errors.append(ManifestError(
                    where, f"a(z) {asset['id']} kompozit konténer `{role}` derivatíva nélkül",
                    'a konténer viszi a futásidejű szöveges sávokat: '
                    '"derivatives": ["captions", "transcript"]'))
        if len(set(components)) != len(components):
            errors.append(ManifestError(
                where, f"a(z) {asset['id']} `composed_of` listája ismétlődik",
                "minden összetevő pontosan egyszer szerepeljen, a lejátszási sorrendben"))
        for cid in components:
            component = assets.get(cid)
            if cid == asset["id"]:
                errors.append(ManifestError(
                    where, f"a(z) {asset['id']} önmagát sorolja összetevőnek", ""))
                continue
            if component is None:
                errors.append(ManifestError(
                    where, f"a(z) {asset['id']} `composed_of` nem létező assetre mutat: {cid}",
                    "javítsd a hivatkozást, vagy vedd fel az összetevő assetet"))
                continue
            if component["file"] != asset["file"]:
                errors.append(ManifestError(
                    where, f"a(z) {asset['id']} összetevője másik fájlban él ({cid})",
                    "a kompozíció egy leckén belüli szerkezet"))
            if component["composed_of"]:
                errors.append(ManifestError(
                    where, f"a(z) {asset['id']} összetevője maga is kompozit ({cid})",
                    "a beágyazott kompozíció nem támogatott — egy szint mély"))
            if not component["source_ref"]:
                errors.append(ManifestError(
                    where, f"a(z) {cid} összetevőnek nincs forrásszövege, így a(z) "
                           f"{asset['id']} szkriptje sem áll össze",
                    "adj `source_ref`-et az összetevőnek"))
            for role in ("captions", "transcript"):
                if role in component["derivatives"]:
                    errors.append(ManifestError(
                        f"{component['file']}:{component['line']}",
                        f"a(z) {cid} `{role}` derivatívát deklarál, de a(z) "
                        f"{asset['id']} konténer része",
                        "a futásidejű szöveges sáv a konténeré — vedd ki a "
                        "derivatívát az összetevőből"))


def _validate_accessibility(assets, errors) -> None:
    """Objective, machine-checkable accessibility structure only.

    Grounded in `02 Tervezet/LMS – hozzáférhetőségi sztenderd.md`:
    synchronised video needs captions (WCAG 2.2 SC 1.2.2) and the transcript does
    not replace them; audio-only needs a full transcript (SC 1.2.1); information-
    carrying visuals need a text equivalent, decorative ones must say so.

    This validates the MANIFEST's structure. Whether the rendered Moodle/H5P page
    conforms remains a release acceptance task, not something a compiler can
    assert.
    """
    components = composed_component_ids(assets)
    for asset in assets.values():
        where = f"{asset['file']}:{asset['line']}"
        kind = asset["kind"]
        a11y = asset["a11y"]
        derivatives = set(asset["derivatives"])
        if asset["mode"] == "reuse":
            continue
        # A component is never delivered on its own: it is cut into the container,
        # which carries the single caption track and transcript for all of them.
        # `_validate_composition` guarantees that container exists and declares
        # both, so the obligation is discharged, not dropped.
        inside_container = asset["id"] in components

        if kind in VISUAL_KINDS or (kind == "video" and a11y.get("visual")):
            visual = a11y.get("visual")
            if visual is None:
                errors.append(ManifestError(
                    where, f"a(z) {asset['id']} vizuális asset, de nincs a11y.visual megjelölve",
                    '"a11y": {"visual": "informative"} vagy {"visual": "decorative"} — '
                    "a dekoratív jelölés is explicit döntés"))
            note_text = " ".join(str(v) for v in a11y.values() if isinstance(v, str))
            if visual == "decorative" and _DEMANDS_ALT.search(f"{asset['title']} {note_text}"):
                errors.append(ManifestError(
                    where, f"a(z) {asset['id']} dekoratívként van jelölve, de a saját "
                           "akadálymentesítési jegyzete alt-szöveget ír elő",
                    'vagy állítsd "informative"-ra és vedd fel az alt-text '
                    "derivatívát, vagy javítsd a jegyzetet"))
            if visual == "informative":
                if "alt-text" not in derivatives:
                    errors.append(ManifestError(
                        where, f"a(z) {asset['id']} tartalmi vizuális, de nincs alt-text derivatívája",
                        'vedd fel: "derivatives": ["alt-text"]'))
                if not a11y.get("alt_source_ref") and not a11y.get("alt_note"):
                    errors.append(ManifestError(
                        where, f"a(z) {asset['id']} tartalmi vizuálishoz nincs alt-szöveg forrás",
                        "adj `a11y.alt_source_ref`-et a leckében lévő @source blokkra, "
                        "vagy `a11y.alt_note`-ot, ha az alt még megírandó"))

        if kind == "video":
            audio = a11y.get("audio")
            if audio is None:
                errors.append(ManifestError(
                    where, f"a(z) {asset['id']} videó, de nincs a11y.audio megjelölve",
                    '"a11y": {"audio": "spoken"} vagy {"audio": "silent"}'))
            elif audio == "spoken" and not inside_container:
                if "captions" not in derivatives:
                    errors.append(ManifestError(
                        where, f"a(z) {asset['id']} beszélt videó felirat-derivatíva nélkül "
                               "(WCAG 2.2 SC 1.2.2)",
                        'vedd fel: "derivatives": ["captions", ...]'))
                if "transcript" not in derivatives:
                    errors.append(ManifestError(
                        where, f"a(z) {asset['id']} beszélt videóhoz nincs leirat-derivatíva",
                        'vedd fel: "derivatives": ["transcript", ...] — '
                        "a projekt szabálya szerint a leirat is kötelező"))

        if kind in AUDIO_KINDS and "transcript" not in derivatives:
            errors.append(ManifestError(
                where, f"a(z) {asset['id']} csak hang, szöveges ekvivalens nélkül "
                       "(WCAG 2.2 SC 1.2.1)",
                'vedd fel: "derivatives": ["transcript"]'))


def _validate_modes(assets, errors) -> None:
    for asset in assets.values():
        where = f"{asset['file']}:{asset['line']}"
        mode = asset["mode"]
        if mode in ("external", "provided"):
            ext = asset["external"]
            if not (ext.get("source") or ext.get("url") or ext.get("path")):
                errors.append(ManifestError(
                    where, f"a(z) {asset['id']} mode={mode}, de nincs forrás-hivatkozása",
                    '"external": {"source": "…"} vagy {"url": "…"} vagy {"path": "…"}'))
        if mode == "human-decision" and not asset["decision"]:
            errors.append(ManifestError(
                where, f"a(z) {asset['id']} emberi döntésre vár, de nincs leírva, mit kell eldönteni",
                '"decision": "<mit kell eldönteni, kinek>" — '
                "eldöntetlen tétel nem tűnhet el csendben"))
        if mode == "generate" and asset["kind"] in ("video", "voiceover") \
                and not asset["source_ref"] and not asset["spec"]:
            errors.append(ManifestError(
                where, f"a(z) {asset['id']} legyártandó hang/videó forrásszöveg és spec nélkül",
                "adj `source_ref`-et a felmondandó szövegre, vagy írj `spec`-et"))


def _validate_sources_used(sources, errors) -> None:
    for src in sources.values():
        if not src["used_by"]:
            errors.append(ManifestError(
                f"{src['file']}:{src['line']}",
                f"a(z) {src['id']} @source blokkra egyetlen asset sem hivatkozik",
                "hivatkozz rá `source_ref`/`a11y.alt_source_ref` mezőből, vagy töröld"))


# ==========================================================================
# Derivative expansion and manifest assembly
# ==========================================================================

#: Production rules whose open question is a RIGHTS question — an AI avatar and
#: voice licence (R2) and image rights / likeness consent (R8). Both are recorded
#: as ⟬KITÖLTENDŐ⟭ organisational decisions; the compiler only reports which
#: assets they hold up, it never resolves them.
#: Structural readiness issues. Unlike R1–R8 these are not organisational
#: decisions waiting to be made — they are facts about the manifest itself, so
#: they outrank every production rule AND an authored `status`. Resolving R3
#: must never turn an asset whose speech has no script into "spec-ready".
MISSING_SPOKEN_SOURCE = "MISSING_SPOKEN_SOURCE"
OPEN_DECISION = "OPEN_DECISION"

READINESS_LABELS = {
    MISSING_SPOKEN_SOURCE: "nincs felmondható forrásszöveg",
    OPEN_DECISION: "nyitott emberi döntés",
}

#: `audio` covers music and sound effects too; only these subtypes are speech.
SPOKEN_AUDIO_SUBTYPES = ("narration", "dialogue")

RIGHTS_BLOCKERS = frozenset({"R2", "R8"})
#: R7 gates on the finalised Moodle course, i.e. on the runtime, not on a rule.
RUNTIME_BLOCKERS = frozenset({"R7"})


def requires_spoken_source(asset: dict) -> bool:
    """Whether producing this asset needs a script that must exist somewhere.

    A voiceover is speech by definition. A video counts when it declares spoken
    audio. An `audio` asset counts only for narration/dialogue — music and SFX
    carry no script, and treating them as missing one would be noise.
    """
    if asset["mode"] != "generate":
        return False
    if asset["kind"] == "voiceover":
        return True
    if asset["kind"] == "audio":
        return asset["subtype"] in SPOKEN_AUDIO_SUBTYPES
    if asset["kind"] == "video":
        return asset["a11y"].get("audio") == "spoken"
    return False


def has_spoken_source(asset: dict) -> bool:
    """Whether this asset's speech resolves to text that lives in the lesson.

    Either its own `@source` block, or — for a composed container — the ordered
    concatenation of its components' blocks. `_validate_composition` refuses a
    component without a source, so a container can never borrow readiness from a
    component that has no script either.
    """
    return bool(asset["source_ref"]) or bool(asset.get("composed_of_resolved"))


def readiness_issues(asset: dict) -> list[str]:
    """Structural reasons this asset cannot be produced, whatever the rules say."""
    issues = []
    if requires_spoken_source(asset) and not has_spoken_source(asset):
        issues.append(MISSING_SPOKEN_SOURCE)
    if asset["decision"]:
        issues.append(OPEN_DECISION)
    return issues


def derive_status(asset: dict) -> str:
    """Status from structural facts first, then authored value, then rules.

    Order matters and is deliberate: a missing script or an unresolved human
    decision is a fact about this asset, so neither an authored `spec-ready` nor
    the resolution of R2/R3/R5/R7/R8 may hide it.
    """
    issues = readiness_issues(asset)
    if MISSING_SPOKEN_SOURCE in issues:
        return "blocked"
    if OPEN_DECISION in issues or asset["mode"] == "human-decision":
        return "pending-human-decision"
    if asset["status"]:
        return asset["status"]
    blockers = set(asset["blockers"])
    if blockers & RIGHTS_BLOCKERS:
        return "pending-rights"
    if blockers & RUNTIME_BLOCKERS:
        return "pending-runtime"
    if blockers:
        return "pending-production-rule"
    return "spec-ready"


def deliverable_id(asset_id: str, role: str | None) -> str:
    if role is None:
        return asset_id
    return f"{asset_id}::{DERIVATIVE_SUFFIX[role]}"


def expand_deliverables(asset: dict, resolved: dict) -> list[dict]:
    """Concrete production outputs for one semantic asset.

    A reuse asset produces nothing new: it points at the canonical asset's
    deliverables. Everything else produces its primary artefact plus one
    deliverable per declared derivative, and derivatives inherit the copy that
    the semantic asset already points at — the author never maintains the same
    narration in five places.
    """
    if asset["mode"] == "reuse":
        return []

    status = derive_status(asset)
    issues = readiness_issues(asset)
    primary_source = resolved.get("primary")
    alt_source = resolved.get("alt")

    out = [{
        "id": deliverable_id(asset["id"], None),
        "asset_id": asset["id"],
        "role": "primary",
        "role_label": "elsődleges",
        "kind": asset["kind"],
        "subtype": asset["subtype"],
        "unit": asset["unit"],
        "module": asset["module"],
        "file": asset["file"],
        "file_kind": asset["file_kind"],
        "title": asset["title"],
        "mode": asset["mode"],
        "status": status,
        "provenance": asset["provenance"],
        "source_id": primary_source["id"] if primary_source else "",
        "text": primary_source["text"] if primary_source else "",
        "text_hash": primary_source["hash"] if primary_source else "",
        "blockers": list(asset["blockers"]),
        "readiness_issues": list(issues),
    }]

    for role in asset["derivatives"]:
        if role in ("captions", "transcript", "voiceover", "audio-only"):
            src = primary_source
        elif role == "alt-text":
            src = alt_source
        else:
            src = None
        out.append({
            "id": deliverable_id(asset["id"], role),
            "asset_id": asset["id"],
            "role": role,
            "role_label": DERIVATIVE_LABELS[role],
            "kind": asset["kind"],
            "subtype": asset["subtype"],
            "unit": asset["unit"],
            "module": asset["module"],
            "file": asset["file"],
            "file_kind": asset["file_kind"],
            "title": f"{asset['title']} — {DERIVATIVE_LABELS[role]}",
            "mode": asset["mode"],
            "status": status,
            "provenance": asset["provenance"],
            "source_id": src["id"] if src else "",
            "text": src["text"] if src else "",
            "text_hash": src["hash"] if src else "",
            "blockers": list(asset["blockers"]),
            "readiness_issues": list(issues),
        })
    return out


SPEC_HASH_FIELDS = (
    "kind", "subtype", "mode", "title", "purpose", "spec", "provenance",
    "provenance_note", "technical", "a11y", "derivatives", "reuse_of",
    "composed_of", "external", "blockers", "production_rules", "decision",
)


def asset_sort_key(asset: dict):
    return (module_sort_index(asset["module"]), file_kind_index(asset["file_kind"]),
            natural_key(asset["unit"]), natural_key(asset["id"]))


def compile_manifest(root: Path = ACTIVE_ROOT, strict: bool = True) -> dict:
    """Discover → parse → validate → resolve → expand. Returns the v2 model."""
    files = discover_sources(root)
    parsed_files = []
    errors: list[ManifestError] = []
    for path in files:
        try:
            parsed_files.append(parse_file(path))
        except ManifestError as exc:
            errors.append(exc)

    _validate_asset_free(parsed_files, errors)
    sources = build_sources(parsed_files, errors)
    assets = build_assets(parsed_files, errors)
    validate(assets, sources, errors)

    if errors and strict:
        raise ManifestValidationFailed(errors)

    ordered_assets = sorted(assets.values(), key=asset_sort_key)
    deliverables: list[dict] = []
    asset_records: list[dict] = []

    for asset in ordered_assets:
        where = f"{asset['file']}:{asset['line']}"
        composed_ids: list[str] = []
        if asset["composed_of"] and not asset["source_ref"]:
            primary_source, primary_text, composed_ids = resolve_composition(
                asset, assets, sources, where)
            asset["composed_of_resolved"] = (
                len(composed_ids) == len(asset["composed_of"]))
        else:
            primary_source, primary_text = (
                resolve_ref(sources, asset["source_ref"], where)
                if asset["source_ref"] else (None, ""))
        alt_source_id = asset["a11y"].get("alt_source_ref", "")
        alt_source, alt_text = (resolve_ref(sources, alt_source_id, where)
                                if alt_source_id else (None, ""))
        resolved = {
            "primary": {**primary_source, "text": primary_text,
                        "hash": sha256_text(primary_text)} if primary_source else None,
            "alt": {**alt_source, "text": alt_text,
                    "hash": sha256_text(alt_text)} if alt_source else None,
        }
        dels = expand_deliverables(asset, resolved)
        deliverables.extend(dels)
        copy_parts = [t for t in (primary_text, alt_text) if t]
        record = dict(asset)
        record["status"] = derive_status(asset)
        record["readiness_issues"] = readiness_issues(asset)
        record["source_text"] = primary_text
        record["source_hash"] = sha256_text(primary_text) if primary_text else ""
        record["source_line"] = primary_source["body_lines"] if primary_source else []
        record["composed_source_ids"] = composed_ids
        record["alt_source_ref"] = alt_source_id
        record["alt_text"] = alt_text
        record["alt_hash"] = sha256_text(alt_text) if alt_text else ""
        record["copy_hash"] = sha256_text("\x1f".join(copy_parts)) if copy_parts else ""
        record["spec_hash"] = sha256_text(
            canonical_json({k: asset[k] for k in SPEC_HASH_FIELDS}))
        record["deliverable_ids"] = [d["id"] for d in dels]
        record["reuse_resolves_to"] = _resolve_reuse(asset, assets)
        asset_records.append(record)

    file_records = []
    for parsed in parsed_files:
        unit, module, file_kind = file_identity(parsed["path"])
        rel = parsed["rel"]
        declared_free = (parsed["asset_free"][0].get("reason", "")
                         if parsed["asset_free"] else "")
        file_records.append({
            "file": rel,
            "unit": unit,
            "module": module,
            "file_kind": file_kind,
            "assets": sorted((a["id"] for a in parsed["assets"] if a.get("id")), key=natural_key),
            "sources": sorted((s["id"] for s in parsed["sources"] if s.get("id")), key=natural_key),
            "asset_free_reason": declared_free,
        })
    file_records.sort(key=lambda f: (module_sort_index(f["module"]),
                                     file_kind_index(f["file_kind"]),
                                     natural_key(f["unit"]), f["file"]))

    return {
        "schema_version": SCHEMA_VERSION,
        "files": file_records,
        "sources": [sources[k] for k in sorted(sources, key=natural_key)],
        "assets": asset_records,
        "deliverables": deliverables,
        "errors": errors,
        "parsed_files": parsed_files,
    }


def resolve_composition(asset: dict, assets: dict[str, dict],
                        sources: dict[str, dict], where: str):
    """Synthetic source for a composed container: its components, in order.

    The text is never copied into the container's own declaration — it is read
    live from the components' `@source` blocks, so editing a scene's narration in
    the lesson moves the container's caption file with it.
    """
    parts, ids = [], []
    for cid in asset["composed_of"]:
        component = assets.get(cid)
        if component is None or not component["source_ref"]:
            continue
        src, text = resolve_ref(sources, component["source_ref"], where)
        if src is None:
            continue
        ids.append(src["id"])
        parts.append(text)
    if not parts:
        return None, "", []
    joined = "\n\n".join(parts)
    return ({"id": " + ".join(ids), "kind": "narration", "body_lines": [],
             "text": joined, "hash": sha256_text(joined)}, joined, ids)


def _resolve_reuse(asset: dict, assets: dict[str, dict]) -> str:
    if asset["mode"] != "reuse":
        return ""
    seen = {asset["id"]}
    current = assets.get(asset["reuse_of"])
    while current is not None and current["mode"] == "reuse" and current["reuse_of"]:
        if current["id"] in seen:
            return ""
        seen.add(current["id"])
        current = assets.get(current["reuse_of"])
    return current["id"] if current else ""


class ManifestValidationFailed(Exception):
    def __init__(self, errors: list[ManifestError]):
        self.errors = errors
        super().__init__(f"{len(errors)} manifest-hiba")


# ==========================================================================
# Statistics — computed once, consumed by every view
# ==========================================================================

def compute_stats(model: dict) -> dict:
    assets = model["assets"]
    deliverables = model["deliverables"]
    files = model["files"]

    def tally(items, key):
        out: dict[str, int] = {}
        for item in items:
            out[key(item)] = out.get(key(item), 0) + 1
        return dict(sorted(out.items(), key=lambda kv: (-kv[1], kv[0])))

    files_with_assets = {a["file"] for a in assets}
    declared_free = sum(1 for f in files if f.get("asset_free_reason"))
    files_with_sources = {s["file"] for s in model["sources"]}
    source_backed = [a for a in assets if a["source_hash"] or a["alt_hash"]]

    return {
        "files_discovered": len(files),
        "files_with_assets": len(files_with_assets),
        "files_asset_free": len(files) - len(files_with_assets),
        "files_declared_asset_free": declared_free,
        "files_with_sources": len(files_with_sources),
        "sources": len(model["sources"]),
        "assets": len(assets),
        "deliverables": len(deliverables),
        "source_backed_assets": len(source_backed),
        "by_mode": tally(assets, lambda a: a["mode"]),
        "by_kind": tally(assets, lambda a: a["kind"]),
        "by_module": {m: sum(1 for a in assets if a["module"] == m)
                      for m in MODULE_ORDER
                      if any(a["module"] == m for a in assets)},
        "by_file_kind": tally(assets, lambda a: a["file_kind"]),
        "by_provenance": tally(assets, lambda a: a["provenance"]),
        "by_status": tally(assets, lambda a: a["status"]),
        "deliverables_by_role": tally(deliverables, lambda d: d["role"]),
        "blockers": tally([b for a in assets for b in a["blockers"]], lambda b: b),
    }


# ==========================================================================
# Renderers
# ==========================================================================

def _write_text(path: Path, text: str) -> bytes:
    data = text.encode("utf-8")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)
    return data


def render_manifest_json(model: dict) -> str:
    payload = {
        "schema_version": model["schema_version"],
        "counts": compute_stats(model),
        "files": model["files"],
        "sources": [{k: v for k, v in s.items() if k != "used_by"} | {"used_by": sorted(s["used_by"])}
                    for s in model["sources"]],
        "assets": [_json_asset(a) for a in model["assets"]],
        "deliverables": model["deliverables"],
    }
    return json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=False) + "\n"


def _json_asset(asset: dict) -> dict:
    keys = ("id", "unit", "module", "file_kind", "file", "line", "kind", "subtype",
            "mode", "status", "title", "purpose", "spec", "provenance", "provenance_note",
            "technical", "a11y", "derivatives", "reuse_of", "reuse_resolves_to", "external",
            "blockers", "production_rules", "readiness_issues", "decision", "notes", "review",
            "source_ref", "composed_of", "composed_source_ids",
            "source_line", "source_text", "source_hash",
            "alt_source_ref", "alt_text", "alt_hash", "copy_hash", "spec_hash",
            "deliverable_ids", "legacy")
    return {k: asset.get(k, "") for k in keys}


ASSET_CSV_HEADER = [
    "ID", "Modul", "Egység", "Fájltípus", "Forrásfájl", "Sor", "Típus", "Altípus",
    "Produkciós mód", "Státusz", "Készültségi akadály", "Cím", "Mit kell gyártani", "Miért (cél)",
    "Forrásblokk", "Felmondandó / generálandó szöveg (élő forrásból)", "Forrás-hash",
    "Alt-forrásblokk", "Alt-szöveg (élő forrásból)", "A11y",
    "Eredet", "Eredet-megjegyzés", "Tech-spec", "Derivatívák", "Deliverable-ek",
    "Újrahasznosítás célja", "Összetevők (kompozit)", "Külső forrás", "Blokkolók",
    "Produkciós szabályok", "Emberi döntés", "Megjegyzés", "Spec-hash", "Régi ID-k",
]


def _flat(value) -> str:
    if isinstance(value, dict):
        return "; ".join(f"{k}: {v}" for k, v in value.items() if v not in ("", None))
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    return "" if value is None else str(value)


def _readiness_cell(record: dict) -> str:
    return "; ".join(READINESS_LABELS.get(issue, issue)
                     for issue in record.get("readiness_issues", []))


def _a11y_cell(asset: dict) -> str:
    a11y = dict(asset["a11y"])
    a11y.pop("alt_source_ref", None)
    return _flat(a11y)


def _source_cell(asset: dict) -> str:
    if asset["source_ref"]:
        return f"`{asset['source_ref']}`"
    if asset["composed_of"]:
        return "kompozit: " + " + ".join(f"`{c}`" for c in asset["composed_of"])
    return "—"


def _legacy_cell(asset: dict) -> str:
    parts = []
    for role in ["asset"] + list(DERIVATIVES):
        ids = asset["legacy"].get(role)
        if ids:
            parts.append(f"{role}: {', '.join(ids)}")
    return " | ".join(parts)


def asset_csv_rows(model: dict) -> list[list[str]]:
    rows = []
    for a in model["assets"]:
        rows.append([
            a["id"], a["module"], a["unit"], a["file_kind"], a["file"], str(a["line"]),
            a["kind"], a["subtype"], MODE_LABELS[a["mode"]], STATUS_LABELS[a["status"]],
            _readiness_cell(a), a["title"], a["spec"], a["purpose"],
            a["source_ref"], a["source_text"], a["source_hash"],
            a["alt_source_ref"], a["alt_text"], _a11y_cell(a),
            PROVENANCE_LABELS[a["provenance"]], a["provenance_note"], _flat(a["technical"]),
            _flat(a["derivatives"]), _flat(a["deliverable_ids"]),
            a["reuse_of"], _flat(a["composed_of"]),
            _flat(a["external"]), _flat(a["blockers"]),
            _flat(a["production_rules"]), a["decision"], a["notes"], a["spec_hash"],
            _legacy_cell(a),
        ])
    return rows


DELIVERABLE_CSV_HEADER = [
    "Deliverable ID", "Asset ID", "Szerep", "Modul", "Egység", "Forrásfájl",
    "Típus", "Altípus", "Produkciós mód", "Státusz", "Készültségi akadály", "Cím",
    "Eredet", "Forrásblokk", "Szöveg (élő forrásból)", "Szöveg-hash", "Blokkolók",
]


def deliverable_csv_rows(model: dict) -> list[list[str]]:
    return [[
        d["id"], d["asset_id"], d["role_label"], d["module"], d["unit"], d["file"],
        d["kind"], d["subtype"], MODE_LABELS[d["mode"]], STATUS_LABELS[d["status"]],
        _readiness_cell(d), d["title"], PROVENANCE_LABELS[d["provenance"]],
        d["source_id"], d["text"], d["text_hash"], _flat(d["blockers"]),
    ] for d in model["deliverables"]]


REUSE_CSV_HEADER = [
    "Újrahasznosító asset", "Modul", "Forrásfájl", "Típus",
    "Kanonikus asset", "Kanonikus fájl", "Kanonikus típus", "Indoklás",
]


def reuse_csv_rows(model: dict) -> list[list[str]]:
    by_id = {a["id"]: a for a in model["assets"]}
    rows = []
    for a in model["assets"]:
        if a["mode"] != "reuse":
            continue
        target = by_id.get(a["reuse_of"], {})
        rows.append([
            a["id"], a["module"], a["file"], a["kind"],
            a["reuse_of"], target.get("file", ""), target.get("kind", ""),
            a["notes"] or a["spec"],
        ])
    return rows


def render_csv(header: list[str], rows: list[list[str]]) -> bytes:
    """UTF-8 with a BOM (so Excel opens Hungarian text correctly) and LF endings.

    Rows end with LF, not CRLF: narration cells legitimately contain their own
    newlines, and a CRLF row terminator made every regeneration trip
    `git diff --check` on trailing whitespace. `.gitattributes` keeps `*.csv
    -text` so the embedded newlines inside quoted cells survive a Windows
    checkout unchanged.
    """
    buf = io.StringIO(newline="")
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow(header)
    writer.writerows(rows)
    return "﻿".encode("utf-8") + buf.getvalue().encode("utf-8")


def _md_cell(value: str) -> str:
    text = "" if value is None else str(value)
    text = text.replace("&amp;", "&")
    text = re.sub(r"\s+", " ", text.replace("|", "∕")).strip()
    return text or "—"


def render_register_md(model: dict) -> str:
    stats = compute_stats(model)
    out: list[str] = []
    P = out.append

    P("# 🎬 Média-asset regiszter — Hasomer Hacair madrichképzés")
    P("")
    P("> **Ez a fájl generált.** Forrás: a `02 Tervezet/` alatti leckefájlok rejtett")
    P("> `@asset` deklarációi és `@source` blokkjai. Ne szerkeszd kézzel —")
    P("> `python3 tools/media_manifest.py build` állítja elő.")
    P(">")
    P("> **Mi ez?** A teljes produkciós leltár: szemantikus assetek (amit a szerző")
    P("> megfogalmaz) és az ezekből származó konkrét deliverable-ek (amit a produkció")
    P("> legyárt) — narráció, videó, animáció, illusztráció, ikon, fotó, hang,")
    P("> felirat, leirat, alt-szöveg, valamint a nyomtatható és letölthető anyagok.")
    P(">")
    P("> A **felmondandó és az alt-szöveg minden buildnél élőben** a leckefájlból")
    P("> származik, nem kifagyasztott pillanatképből.")
    P("")
    P("## 📊 Összesítő")
    P("")
    P("| Mutató | Érték |")
    P("|---|--:|")
    P(f"| Feldolgozott forrásfájl | **{stats['files_discovered']}** |")
    P(f"| Assetet tartalmazó fájl | {stats['files_with_assets']} |")
    P(f"| Ellenőrzötten asset nélküli fájl | {stats['files_asset_free']} |")
    P(f"| Forrásblokk (`@source`) | {stats['sources']} |")
    P(f"| Szemantikus asset | **{stats['assets']}** |")
    P(f"| Produkciós deliverable | **{stats['deliverables']}** |")
    P("")
    P("**Produkciós mód szerint**")
    P("")
    P("| Mód | Db |")
    P("|---|--:|")
    for mode in MODES:
        if mode in stats["by_mode"]:
            P(f"| {MODE_LABELS[mode]} | {stats['by_mode'][mode]} |")
    P("")
    P("**Asset-típus szerint**")
    P("")
    P("| Típus | Db |")
    P("|---|--:|")
    for kind, count in stats["by_kind"].items():
        P(f"| {kind} | {count} |")
    P("")
    P("**Deliverable-szerep szerint**")
    P("")
    P("| Szerep | Db |")
    P("|---|--:|")
    for role, count in stats["deliverables_by_role"].items():
        label = "elsődleges" if role == "primary" else DERIVATIVE_LABELS.get(role, role)
        P(f"| {label} | {count} |")
    P("")
    P("**Modul szerint**")
    P("")
    P("| Modul | Db |")
    P("|---|--:|")
    for module, count in stats["by_module"].items():
        P(f"| {module} | {count} |")
    P("")
    P("**Státusz szerint**")
    P("")
    P("| Státusz | Db |")
    P("|---|--:|")
    for status, count in stats["by_status"].items():
        P(f"| {STATUS_LABELS[status]} | {count} |")
    P("")

    if stats["blockers"]:
        P("**Nyitott produkciós blokkolók (hivatkozások szerint)**")
        P("")
        P("| Blokkoló | Érintett asset |")
        P("|---|--:|")
        for blocker, count in stats["blockers"].items():
            P(f"| {blocker} | {count} |")
        P("")

    open_rules = [r for r in production_rules() if "KITÖLTENDŐ" in r["text"]]
    if open_rules:
        P("## ⛔ Nyitott produkciós kapuk")
        P("")
        P("Ezek **szervezeti és jogi döntések**. Amíg nyitva vannak, a jelölt")
        P("assetek kötegelt gyártása nem indulhat. A jelölés gépileg is")
        P("detektálható, ezért a `content_integrity.py --release-report` számolja.")
        P("")
        P("| Szabály | Mi hiányzik | Érintett asset |")
        P("|---|---|--:|")
        for rule in open_rules:
            missing = re.search(r"([^.;]*⟬\s*KITÖLTENDŐ\s*⟭[^.;]*)", rule["text"])
            text = _md_cell(missing.group(1).strip() if missing else "⟬KITÖLTENDŐ⟭")
            P(f"| **{rule['id']}** — {_md_cell(rule['title'])} | {text} "
              f"| {stats['blockers'].get(rule['id'], 0)} |")
        P("")

    P("## 🗂 Assetek fájlonként")
    P("")
    by_file: dict[str, list[dict]] = {}
    for asset in model["assets"]:
        by_file.setdefault(asset["file"], []).append(asset)
    file_order = [f for f in model["files"] if f["file"] in by_file]
    for file_rec in file_order:
        rel = file_rec["file"]
        P(f"### {rel}")
        P("")
        P(f"*Egység:* `{file_rec['unit']}` · *típus:* {file_rec['file_kind']}")
        P("")
        P("| ID | Típus | Mód | Státusz | Cím | Forrásblokk | Derivatívák | Eredet |")
        P("|---|---|---|---|---|---|---|---|")
        for asset in by_file[rel]:
            subtype = f"/{asset['subtype']}" if asset["subtype"] else ""
            P("| `{id}` | {kind}{sub} | {mode} | {status} | {title} | {src} | {der} | {prov} |".format(
                id=asset["id"], kind=asset["kind"], sub=subtype,
                mode=MODE_LABELS[asset["mode"]], status=STATUS_LABELS[asset["status"]],
                title=_md_cell(asset["title"]),
                src=_source_cell(asset),
                der=_md_cell(", ".join(DERIVATIVE_LABELS[d] for d in asset["derivatives"])),
                prov=PROVENANCE_LABELS[asset["provenance"]]))
        P("")

    asset_free = [f for f in model["files"] if not f["assets"]]
    if asset_free:
        P("## 🈳 Ellenőrzötten asset nélküli fájlok")
        P("")
        P("| Fájl | Egység | Típus | Indoklás |")
        P("|---|---|---|---|")
        for file_rec in asset_free:
            P(f"| {file_rec['file']} | `{file_rec['unit']}` | {file_rec['file_kind']} "
              f"| {_md_cell(file_rec.get('asset_free_reason', ''))} |")
        P("")

    blocked = [a for a in model["assets"] if a["readiness_issues"]]
    if blocked:
        P("## ⛔ Készültségi akadályok")
        P("")
        P("Ezek **nem** produkciós szabályra várnak: hiányzik valami, ami nélkül az")
        P("asset egyáltalán nem gyártható. Egy R2/R3/R5 feloldása sem teszi őket")
        P("készre, és kézzel beírt `status` sem írja felül.")
        P("")
        P("| ID | Fájl | Akadály | Státusz |")
        P("|---|---|---|---|")
        for asset in blocked:
            P(f"| `{asset['id']}` | {asset['file']} | {_readiness_cell(asset)} "
              f"| {STATUS_LABELS[asset['status']]} |")
        P("")

    human = [a for a in model["assets"]
             if a["mode"] == "human-decision" or a["decision"]]
    if human:
        P("## ⚖️ Emberi döntésre váró tételek")
        P("")
        P("| ID | Fájl | Mit kell eldönteni |")
        P("|---|---|---|")
        for asset in human:
            P(f"| `{asset['id']}` | {asset['file']} | {_md_cell(asset['decision'])} |")
        P("")

    P("---")
    P("")
    P("Újraépítés: `python3 tools/media_manifest.py build` ·")
    P("ellenőrzés: `python3 tools/media_manifest.py check`")
    return "\n".join(out) + "\n"


# --------------------------------------------------------------------------
# XLSX
# --------------------------------------------------------------------------

def render_xlsx(model: dict) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - environment dependent
        raise ManifestError("xlsx", f"openpyxl nem elérhető: {exc}",
                            "pip install openpyxl, vagy hagyd ki a workbook-építést")

    stats = compute_stats(model)
    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    mode_fill = {
        "generate": PatternFill("solid", fgColor="E2EFDA"),
        "reuse": PatternFill("solid", fgColor="FFF2CC"),
        "external": PatternFill("solid", fgColor="DEEBF7"),
        "provided": PatternFill("solid", fgColor="DEEBF7"),
        "human-decision": PatternFill("solid", fgColor="F8CBAD"),
    }

    def sheet(ws, headers, widths, wrapcols):
        for col, (head, width) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(1, col, head)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        ws.row_dimensions[1].height = 30

    def fill_rows(ws, rows, wrapcols, tint=None):
        for r, values in enumerate(rows, 2):
            for c, value in enumerate(values, 1):
                cell = ws.cell(r, c, value)
                cell.border = border
                cell.alignment = wrap if c in wrapcols else top
            if tint:
                key, col = tint
                fill = mode_fill.get(key(rows[r - 2]))
                if fill:
                    ws.cell(r, col).fill = fill

    ws = wb.active
    ws.title = "Assetek"
    widths = [16, 7, 10, 12, 34, 7, 14, 16, 18, 22, 26, 26, 46, 34, 16, 72, 18,
              16, 46, 30, 16, 26, 26, 24, 34, 16, 30, 26, 16, 18, 30, 26, 18, 30]
    wrapcols = {11, 12, 13, 14, 16, 19, 20, 22, 23, 24, 25, 26, 28, 31, 32, 34}
    sheet(ws, ASSET_CSV_HEADER, widths, wrapcols)
    rows = asset_csv_rows(model)
    modes = [a["mode"] for a in model["assets"]]
    for r, values in enumerate(rows, 2):
        for c, value in enumerate(values, 1):
            cell = ws.cell(r, c, value)
            cell.border = border
            cell.alignment = wrap if c in wrapcols else top
        fill = mode_fill.get(modes[r - 2])
        if fill:
            ws.cell(r, 9).fill = fill
        if rows[r - 2][10]:
            ws.cell(r, 11).fill = mode_fill["human-decision"]

    ws2 = wb.create_sheet("Deliverable-ek")
    sheet(ws2, DELIVERABLE_CSV_HEADER,
          [24, 16, 20, 7, 10, 34, 14, 16, 18, 22, 26, 34, 16, 16, 72, 18, 20],
          {11, 12, 15, 17})
    fill_rows(ws2, deliverable_csv_rows(model), {11, 12, 15, 17})

    ws3 = wb.create_sheet("Újrahasznosítás")
    sheet(ws3, REUSE_CSV_HEADER, [18, 7, 34, 14, 18, 34, 14, 70], {3, 6, 8})
    fill_rows(ws3, reuse_csv_rows(model), {3, 6, 8})

    # Two different questions, two sheets. Mixing them buried the single asset
    # that genuinely needs a human answer under several hundred rule-blocked rows.
    ws4 = wb.create_sheet("Emberi döntések")
    sheet(ws4, ["ID", "Modul", "Fájl", "Típus", "Mód", "Mit kell eldönteni"],
          [18, 7, 34, 14, 20, 100], {3, 6})
    fill_rows(ws4, [[a["id"], a["module"], a["file"], a["kind"],
                     MODE_LABELS[a["mode"]], a["decision"]]
                    for a in model["assets"]
                    if a["mode"] == "human-decision" or a["decision"]], {3, 6})

    ws4b = wb.create_sheet("Blokkolt assetek")
    sheet(ws4b, ["ID", "Modul", "Fájl", "Típus", "Státusz", "Blokkolók"],
          [18, 7, 34, 14, 24, 24], {3, 6})
    fill_rows(ws4b, [[a["id"], a["module"], a["file"], a["kind"],
                      STATUS_LABELS[a["status"]], _flat(a["blockers"])]
                     for a in model["assets"] if a["blockers"]], {3, 6})

    ws5 = wb.create_sheet("Migráció")
    recon = model.get("reconciliation")
    sheet(ws5, MIGRATION_CSV_HEADER, [18, 20, 34, 20, 16, 20, 26, 22, 60], {3, 9})
    fill_rows(ws5, recon["rows"] if recon else [], {3, 9})

    ws6 = wb.create_sheet("Produkciós konvenciók")
    sheet(ws6, ["Szabály", "Megnevezés", "Tartalom"], [10, 30, 120], {2, 3})
    fill_rows(ws6, [[r["id"], r["title"], r["text"]] for r in production_rules()], {2, 3})

    ws7 = wb.create_sheet("Asset nélküli fájlok")
    sheet(ws7, ["Fájl", "Egység", "Típus", "Indoklás"], [70, 14, 16, 80], {1, 4})
    fill_rows(ws7, [[f["file"], f["unit"], f["file_kind"], f.get("asset_free_reason", "")]
                    for f in model["files"] if not f["assets"]], {1, 4})

    ws0 = wb.create_sheet("Összesítő")
    ws0.sheet_view.showGridLines = False
    ws0.column_dimensions["A"].width = 46
    ws0.column_dimensions["B"].width = 12
    title_font = Font(bold=True, size=13)
    sub_font = Font(bold=True, size=11, color="1F4E78")
    ws0["A1"] = "Média-asset regiszter v2 — összesítő"
    ws0["A1"].font = title_font
    row = 3
    for label, value in (
        ("Feldolgozott forrásfájl", stats["files_discovered"]),
        ("Assetet tartalmazó fájl", stats["files_with_assets"]),
        ("Asset nélküli fájl", stats["files_asset_free"]),
        ("Forrásblokk (@source)", stats["sources"]),
        ("Szemantikus asset", stats["assets"]),
        ("Produkciós deliverable", stats["deliverables"]),
        ("Forrás-szöveggel kötött asset", stats["source_backed_assets"]),
    ):
        ws0.cell(row, 1, label)
        ws0.cell(row, 2, value)
        row += 1
    row += 1

    def block(start, heading, mapping, labels=None):
        ws0.cell(start, 1, heading).font = sub_font
        cursor = start + 1
        for key, value in mapping.items():
            ws0.cell(cursor, 1, labels.get(key, key) if labels else key)
            ws0.cell(cursor, 2, value)
            cursor += 1
        return cursor + 1

    row = block(row, "Produkciós mód szerint", stats["by_mode"], MODE_LABELS)
    row = block(row, "Asset-típus szerint", stats["by_kind"])
    row = block(row, "Modul szerint", stats["by_module"])
    row = block(row, "Eredet szerint", stats["by_provenance"], PROVENANCE_LABELS)
    row = block(row, "Státusz szerint", stats["by_status"], STATUS_LABELS)
    row = block(row, "Deliverable-szerep szerint", stats["deliverables_by_role"],
                {"primary": "elsődleges", **DERIVATIVE_LABELS})
    if stats["blockers"]:
        row = block(row, "Blokkolók szerint", stats["blockers"])

    wb.move_sheet("Összesítő", -(len(wb.sheetnames) - 1))

    # openpyxl stamps docProps with the current time and the zip entries with the
    # current clock. Both are pinned so the same tree always yields the same bytes
    # and `check` never reports a phantom difference.
    wb.properties.creator = "tools/media_manifest.py"
    wb.properties.lastModifiedBy = "tools/media_manifest.py"
    wb.properties.created = FIXED_TIMESTAMP
    wb.properties.modified = FIXED_TIMESTAMP

    raw = io.BytesIO()
    wb.save(raw)
    return _normalise_zip(raw.getvalue())


import datetime as _dt  # noqa: E402  (kept next to its single use)

PRODUCTION_RULES_FILE = MEDIA_ROOT / "produkcios-szabalyok.json"

#: Fixed timestamp for workbook metadata and zip entries. Not a build date — a
#: constant, so two builds of the same tree are byte-identical.
FIXED_TIMESTAMP = _dt.datetime(2020, 1, 1, 0, 0, 0)


#: openpyxl stamps `dcterms:modified` with the wall clock at save time, after the
#: workbook properties are set, so it has to be pinned in the produced bytes.
_MODIFIED_RE = re.compile(
    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)")
FIXED_ISO = b"2020-01-01T00:00:00Z"


def _normalise_zip(data: bytes) -> bytes:
    """Rewrite an xlsx with sorted entries, constant timestamps and no clock."""
    out = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(data)) as src:
        names = sorted(src.namelist())
        with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as dst:
            for name in names:
                payload = src.read(name)
                if name == "docProps/core.xml":
                    payload = _MODIFIED_RE.sub(rb"\g<1>" + FIXED_ISO + rb"\g<2>", payload)
                info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
                info.compress_type = zipfile.ZIP_DEFLATED
                info.external_attr = 0o600 << 16
                info.create_system = 0
                dst.writestr(info, payload)
    return out.getvalue()


# ==========================================================================
# Production rules (organisational conventions carried forward from v1)
# ==========================================================================

_PRODUCTION_RULES_CACHE: list[dict] | None = None


def production_rules() -> list[dict]:
    """R1–R8 — organisational conventions, canonical in their own data file.

    They are not compiler logic: v2 references them by ID from `blockers` /
    `production_rules` instead of copying their text onto hundreds of rows. The
    text lives in `produkcios-szabalyok.json` rather than in the retired v1
    snapshot, so resolving R2/R3/R5 does not mean editing frozen forensic data.
    A missing file is a hard error — a silently empty conventions sheet would be
    worse than a failed build.
    """
    global _PRODUCTION_RULES_CACHE
    if _PRODUCTION_RULES_CACHE is None:
        if not PRODUCTION_RULES_FILE.exists():
            raise ManifestError(
                _rel(PRODUCTION_RULES_FILE), "hiányzik a produkciós konvenciók adatfájlja",
                "állítsd vissza a fájlt — a fordító nem a befagyasztott v1 JSON-ból dolgozik")
        data = json.loads(PRODUCTION_RULES_FILE.read_text(encoding="utf-8"))
        _PRODUCTION_RULES_CACHE = data["rules"]
    return _PRODUCTION_RULES_CACHE


# ==========================================================================
# Legacy reconciliation
# ==========================================================================

MIGRATION_CSV_HEADER = [
    "old_id", "old_type", "old_file", "old_category", "old_generate_state",
    "v2_asset_id", "v2_deliverable_id", "status", "reason",
]

RECON_STATUSES = (
    "PRESERVED", "CHANGED", "DERIVED_NOW", "MERGED_INTO_PARENT", "REUSE",
    "REPLACED", "REMOVED_STALE", "NO_LONGER_REQUIRED", "AMBIGUOUS",
    "CURRENTLY_UNMAPPED_ERROR",
)


def load_legacy() -> dict | None:
    for candidate in (LEGACY_JSON, LEGACY_JSON_FALLBACK):
        if candidate.exists():
            return json.loads(candidate.read_text(encoding="utf-8"))
    return None


def load_dispositions() -> dict[str, dict]:
    """Hand-authored disposition for legacy rows with no v2 home.

    Only rows that genuinely have no current requirement land here, each with a
    reason. It is historical reconciliation input, never current copy.
    """
    if not LEGACY_DISPOSITIONS.exists():
        return {}
    data = json.loads(LEGACY_DISPOSITIONS.read_text(encoding="utf-8"))
    return {entry["old_id"]: entry for entry in data.get("rows", [])}


def reconcile(model: dict) -> dict:
    legacy = load_legacy()
    if legacy is None:
        raise ManifestError("reconcile", "a történeti media-merged.json nem található",
                            f"várt helyek: {_rel(LEGACY_JSON)} vagy {_rel(LEGACY_JSON_FALLBACK)}")
    dispositions = load_dispositions()
    legacy_rows = legacy["assets"]

    #: old id -> (v2 asset, role)
    claimed: dict[str, list[tuple[dict, str]]] = {}
    for asset in model["assets"]:
        for role, ids in asset["legacy"].items():
            for old_id in ids:
                claimed.setdefault(old_id, []).append((asset, role))

    rows: list[list[str]] = []
    tally: dict[str, int] = {status: 0 for status in RECON_STATUSES}
    conflicts: list[str] = []

    for legacy_row in sorted(legacy_rows, key=lambda r: natural_key(r["assetId"])):
        old_id = legacy_row["assetId"]
        claims = claimed.get(old_id, [])
        if len(claims) > 1:
            conflicts.append(
                f"{old_id}: több v2 asset is igényli "
                f"({', '.join(a['id'] + ':' + role for a, role in claims)})")
        if claims:
            asset, role = claims[0]
            status, deliverable, reason = _classify_claim(asset, role, legacy_row)
        elif old_id in dispositions:
            entry = dispositions[old_id]
            status = entry["status"]
            deliverable = entry.get("v2_deliverable_id", "")
            reason = entry["reason"]
            if status not in RECON_STATUSES:
                conflicts.append(f"{old_id}: ismeretlen diszpozíció-státusz {status!r}")
            asset = None
        else:
            status = "CURRENTLY_UNMAPPED_ERROR"
            deliverable = ""
            reason = ("nincs v2 megfelelője és nincs dokumentált diszpozíciója sem — "
                      "vedd fel a legacy blokkba vagy a legacy-dispositions.json-be")
            asset = None
        tally[status] = tally.get(status, 0) + 1
        rows.append([
            old_id, legacy_row["assetType"], legacy_row["file"], legacy_row["category"],
            legacy_row.get("dedup") and "dedup-tag" or "produce",
            asset["id"] if asset else (dispositions.get(old_id, {}).get("v2_asset_id", "")),
            deliverable, status, reason,
        ])

    return {
        "rows": rows,
        "tally": tally,
        "legacy_total": len(legacy_rows),
        "mapped_total": sum(tally.values()),
        "unmapped": tally.get("CURRENTLY_UNMAPPED_ERROR", 0),
        "conflicts": conflicts,
        "legacy_dedup_groups": sum(len(g.get("groups", [])) for g in legacy.get("dedup", [])),
    }


def _classify_claim(asset: dict, role: str, legacy_row: dict) -> tuple[str, str, str]:
    if role == "asset":
        if asset["mode"] == "reuse":
            return ("REUSE", "",
                    f"explicit újrahasznosítás: {asset['id']} → {asset['reuse_of']}")
        if asset["mode"] == "human-decision":
            return ("AMBIGUOUS", deliverable_id(asset["id"], None),
                    f"emberi döntésre vár: {asset['decision'][:160]}")
        if asset["id"] == legacy_row["assetId"]:
            return ("PRESERVED", deliverable_id(asset["id"], None),
                    "azonos azonosítójú szemantikus asset a jelenlegi forrásban")
        return ("CHANGED", deliverable_id(asset["id"], None),
                f"azonosító a v2 egység-névtérre változott ({legacy_row['assetId']} → {asset['id']})")
    if asset["mode"] == "reuse":
        return ("REUSE", "",
                f"a szülő asset újrahasznosítás: {asset['id']} → {asset['reuse_of']}")
    if role in asset["derivatives"]:
        return ("DERIVED_NOW", deliverable_id(asset["id"], role),
                f"a v2-ben a(z) {asset['id']} `{role}` derivatívája, "
                "a szövegét a forrásblokk adja")
    return ("MERGED_INTO_PARENT", deliverable_id(asset["id"], None),
            f"a v2-ben a(z) {asset['id']} deklarációjának része "
            f"(nincs külön legyártandó `{role}` deliverable)")


def render_migration_md(model: dict, recon: dict) -> str:
    stats = compute_stats(model)
    legacy = load_legacy() or {}
    legacy_rows = legacy.get("assets", [])
    legacy_reuse = sum(1 for r in legacy_rows if r.get("dedup"))
    out: list[str] = []
    P = out.append

    P("# Asset Manifest v2 — migrációs és egyeztetési jelentés")
    P("")
    P("> **Generált fájl.** `python3 tools/media_manifest.py build` állítja elő a")
    P("> jelenlegi leckefájlok `@asset`/`@source` deklarációiból és a befagyasztott")
    P("> történeti leltárból. Ne szerkeszd kézzel.")
    P("")
    P("## 1. Mi változott az architektúrában")
    P("")
    P("| | Régi (v1) | Új (v2) |")
    P("|---|---|---|")
    P("| Kánoni forrás | befagyasztott `media-merged.json` | a jelenlegi Markdown `@asset`/`@source` deklarációi |")
    P("| Szövegek | kinyeréskori pillanatkép (`verbatim`) | minden buildnél élőben a leckéből |")
    P("| Helyhivatkozás | sorszám-pillanatkép (`lineRef`) | stabil forrásblokk-ID |")
    P("| Kinyerés | AI-workflow (nyugdíjazott) | determinisztikus fordító, hálózat és AI nélkül |")
    P("| Újrahasznosítás | utólagos dedup-elemzés | explicit `mode: reuse` + `reuse_of` |")
    P("| Elcsúszás észlelése | nincs | `media_manifest.py check` a CI-ban |")
    P("")
    P("## 2. Történeti alap")
    P("")
    P(f"- **{len(legacy_rows)} sor** a befagyasztott leltárban")
    P(f"- ebből **{len(legacy_rows) - legacy_reuse} legyártandó** és **{legacy_reuse} újrahasznosítás** "
      "(a régi `Legyártandó?` besorolás szerint)")
    P(f"- **{recon['legacy_dedup_groups']} dedup-csoport**")
    P("")
    P("## 3. Jelenlegi v2 leltár")
    P("")
    P("| Mutató | Érték |")
    P("|---|--:|")
    P(f"| Szemantikus asset | **{stats['assets']}** |")
    P(f"| Produkciós deliverable | **{stats['deliverables']}** |")
    for mode in MODES:
        if mode in stats["by_mode"]:
            P(f"| ebből {MODE_LABELS[mode]} | {stats['by_mode'][mode]} |")
    P(f"| Forrásblokk | {stats['sources']} |")
    P(f"| Feldolgozott fájl | {stats['files_discovered']} |")
    P(f"| Assetet tartalmazó fájl | {stats['files_with_assets']} |")
    P("")
    P("A két szám **nem összemérhető közvetlenül**: a régi leltár egy táblában")
    P("keverte a szemantikus követelményt és a belőle származó akadálymentesítési")
    P("deliverable-t (felirat, leirat, alt-szöveg külön sorként). A v2 ezeket")
    P("derivatívaként a szülő assethez köti, ezért kevesebb *asset* és külön")
    P("számolt *deliverable* keletkezik.")
    P("")
    P("## 4. A 747 történeti sor egyeztetése")
    P("")
    P("| Diszpozíció | Db | Jelentése |")
    P("|---|--:|---|")
    meanings = {
        "PRESERVED": "azonos azonosítóval megmaradt szemantikus asset",
        "CHANGED": "megmaradt, de a v2 egység-névtérben új azonosítót kapott",
        "DERIVED_NOW": "a v2-ben egy szülő asset derivatívája (felirat / leirat / alt / hang)",
        "MERGED_INTO_PARENT": "a szülő asset deklarációjába olvadt, nincs külön deliverable",
        "REUSE": "explicit újrahasznosítás egy kanonikus assetre",
        "REPLACED": "más asset váltotta ki",
        "REMOVED_STALE": "a jelenlegi forrásban nincs ilyen követelmény",
        "NO_LONGER_REQUIRED": "a jelenlegi tananyag már nem igényli",
        "AMBIGUOUS": "emberi döntés kell a besoroláshoz",
        "CURRENTLY_UNMAPPED_ERROR": "**hiba** — egyetlen sor sem maradhat itt",
    }
    for status in RECON_STATUSES:
        count = recon["tally"].get(status, 0)
        if count or status == "CURRENTLY_UNMAPPED_ERROR":
            P(f"| `{status}` | {count} | {meanings[status]} |")
    P("")
    P(f"**Összesen: {recon['mapped_total']} / {recon['legacy_total']} sor diszpozícionálva.**")
    P(f"Nem egyeztetett (hiba): **{recon['unmapped']}**.")
    P("")
    if recon["conflicts"]:
        P("### Ütközések")
        P("")
        for conflict in recon["conflicts"]:
            P(f"- {conflict}")
        P("")
    P("A soronkénti leképezés gépi formában: `asset-migration-map.csv`.")
    P("")
    unsourced = [a for a in model["assets"]
                 if a["mode"] == "generate" and not has_spoken_source(a)
                 and ({"captions", "transcript"} & set(a["derivatives"]))]
    if unsourced:
        P("## 5. Felirat/leirat forrásszöveg nélkül")
        P("")
        P("Ezeknél a beszélt asseteknél a jelenlegi lecke nem tartalmaz olyan")
        P("összefüggő, idézett szkriptet, amit `@source` blokkba lehetett volna fogni,")
        P("ezért a felirat- és leirat-deliverable **szöveg nélkül** áll. A produkció")
        P("nem indulhat el rajtuk, amíg a szkript be nem kerül a leckébe — utána a")
        P("`@source` blokk és a `source_ref` felvételével a szöveg automatikusan")
        P("bekerül a regiszterbe.")
        P("")
        P("| ID | Fájl | Derivatívák |")
        P("|---|---|---|")
        for asset in unsourced:
            P(f"| `{asset['id']}` | {asset['file']} | "
              f"{', '.join(asset['derivatives'])} |")
        P("")

    P("## 6. Nyitott produkciós döntések")
    P("")
    P("| ID | Fájl | Mit kell eldönteni |")
    P("|---|---|---|")
    for asset in model["assets"]:
        if asset["mode"] == "human-decision" or asset["decision"]:
            P(f"| `{asset['id']}` | {asset['file']} | {_md_cell(asset['decision'])} |")
    P("")
    if stats["blockers"]:
        P("Hivatkozott produkciós szabályok / blokkolók:")
        P("")
        for blocker, count in stats["blockers"].items():
            P(f"- **{blocker}** — {count} asset")
        P("")
    return "\n".join(out) + "\n"


# ==========================================================================
# Discovery lint
# ==========================================================================

#: (regex, why, confidence). Confidence HIGH findings must be resolved; the rest
#: stay warnings so CI does not drown in keyword noise.
LINT_SIGNALS: tuple[tuple[str, str, str], ...] = (
    (r"\bNarráció\b|\bMit hallunk\b|\bvoice-?over\b|\bfelmondandó\b",
     "narráció / felmondandó szöveg", "HIGH"),
    (r"\bAI beszélő fej\b|\bbeszélő fej\b|\bavatar\b", "beszélőfej-videó", "HIGH"),
    # A video or an audio clip always drags a mandatory text equivalent with it,
    # so an undeclared one is never a low-stakes omission.
    (r"\bvideó\b|\bvideo\b|\bInteractive Video\b", "videó", "HIGH"),
    (r"\bhang(?:sáv|felvétel|alámondás)\b|\bpodcast\b", "hang", "HIGH"),
    (r"\banimáció\b|\banimált\b", "animáció", "MEDIUM"),
    (r"\billusztráció\b|\bgrafika\b", "illusztráció / grafika", "MEDIUM"),
    (r"\bdiagram\b|\bábra\b", "diagram / ábra", "MEDIUM"),
    (r"\bikon\b", "ikon", "MEDIUM"),
    (r"\bfotó\b|\bstock\b|\bképernyőkép\b|\bscreenshot\b", "fotó / képernyőkép", "MEDIUM"),
    (r"\bzene\b|\bSFX\b|\bhangeffekt\b", "zene / SFX", "MEDIUM"),
    (r"\bfelirat\b|\bleirat\b|\btranszkript\b", "felirat / leirat", "MEDIUM"),
    (r"\balt-szöveg\b|\balt szöveg\b", "alt-szöveg", "HIGH"),
    (r"\bnyomtat\w*\b|\bkinyomtat\w*\b", "nyomtatható anyag", "HIGH"),
    (r"\bmunkalap\b|\bworksheet\b", "munkalap", "HIGH"),
    (r"\bkártya\b|\bkártyapakli\b|\bkártyaszett\b", "kártyaszett", "HIGH"),
    (r"\bposzter\b|\bplakát\b|\bflipchart-sablon\b", "poszter", "HIGH"),
    (r"\bsablon\b|\bhandout\b|\brésztvevői lap\b", "sablon / handout", "MEDIUM"),
    (r"\bletölthető\b|\bletöltés\b|\bPDF\b|\bDOCX\b", "letölthető dokumentum", "HIGH"),
    (r"\bellenőrzőlista\b|\bchecklist\b", "ellenőrzőlista", "MEDIUM"),
    (r"\btanúsítvány\b|\bcertificate\b", "tanúsítvány", "MEDIUM"),
)

_LINT_COMPILED = tuple((re.compile(pattern, re.I), why, conf)
                       for pattern, why, conf in LINT_SIGNALS)

_HEADING_RE = re.compile(r"^(#{1,6})\s+(.*?)\s*$")


def lint(model: dict) -> list[dict]:
    """Conservative, deterministic search for asset requirements without a declaration.

    Section-scoped: a signal is only reported when the section that contains it —
    or its enclosing slide/block section — has no `@asset` declaration at all. A
    narration marker whose blockquote is not inside a `@source` block is always
    reported, because that is the exact failure the v1 architecture could not see.
    """
    findings: list[dict] = []
    for parsed in model["parsed_files"]:
        if not in_lint_scope(parsed["path"]) or parsed["asset_free"]:
            continue
        lines = parsed["lines"]
        covered = parsed["covered"]

        def is_metadata(idx1: int) -> bool:
            return any(start <= idx1 <= end for start, end in covered)

        sections = _sections(lines)
        decl_lines = [a["_line"] for a in parsed["assets"]]
        src_lines = [(s["_body_lines"][0], s["_body_lines"][1]) for s in parsed["sources"]]

        for start, end, level, title in sections:
            has_decl = any(start <= line <= end for line in decl_lines)
            covering = [s for s in sections if s[0] <= start and s[1] >= end and s[2] < level]
            parent_decl = any(any(p[0] <= line <= p[1] for line in decl_lines) for p in covering)
            for idx in range(start, min(end, len(lines)) + 1):
                if idx > len(lines):
                    break
                line = lines[idx - 1]
                if is_metadata(idx) or not line.strip():
                    continue
                # A table row is tabular content; its column headers ("| Kártya |")
                # are labels inside an answer key, not production requirements.
                if line.lstrip().startswith("|"):
                    continue
                for pattern, why, confidence in _LINT_COMPILED:
                    if not pattern.search(line):
                        continue
                    if why == "narráció / felmondandó szöveg":
                        # Only a label or heading introduces a script. The same word
                        # inside a blockquote is the accessibility note talking
                        # ABOUT the narration.
                        if line.lstrip().startswith(">"):
                            continue
                        quote = _script_after(lines, idx)
                        if quote:
                            covered_by_source = any(a <= quote <= b for a, b in src_lines)
                            if not covered_by_source:
                                findings.append(_finding(
                                    parsed["rel"], idx, title,
                                    "HIGH", "forrásblokk nélküli narráció-idézet",
                                    "tedd @source blokkba, és hivatkozz rá `source_ref`-fel"))
                            continue
                    if has_decl or parent_decl:
                        continue
                    findings.append(_finding(parsed["rel"], idx, title, confidence, why,
                                             "vedd fel @asset deklarációval, "
                                             "vagy igazold, hogy nem produkciós követelmény"))
    seen = set()
    unique = []
    for finding in findings:
        key = (finding["file"], finding["line"], finding["reason"])
        if key in seen:
            continue
        seen.add(key)
        unique.append(finding)
    unique.sort(key=lambda f: (f["confidence"] != "HIGH", f["file"], f["line"]))
    return unique


def _finding(rel, line, heading, confidence, reason, fix) -> dict:
    return {"file": rel, "line": line, "heading": heading,
            "confidence": confidence, "reason": reason, "fix": fix}


def _sections(lines: list[str]) -> list[tuple[int, int, int, str]]:
    heads = []
    in_fence = False
    for idx, line in enumerate(lines, 1):
        if line.lstrip().startswith("```"):
            in_fence = not in_fence
            continue
        if in_fence:
            continue
        match = _HEADING_RE.match(line)
        if match:
            heads.append((idx, len(match.group(1)), match.group(2)))
    sections = []
    for pos, (start, level, title) in enumerate(heads):
        end = len(lines)
        for later_start, later_level, _ in heads[pos + 1:]:
            if later_level <= level:
                end = later_start - 1
                break
        sections.append((start, end, level, title))
    return sections


#: Between a narration label and the script itself the corpus places an italic
#: tone direction and/or a ♿ accessibility reminder — sometimes as its own
#: blockquote. Those are instructions about the narration, so the lint has to
#: look past them exactly as the source blocks do.
_A11Y_NOTE = re.compile(r"♿|akadálymentes|wcag|felirat \+ teljes", re.I)
_STAGE_DIRECTION = re.compile(r"^\*\(.*\)\*$")


def _script_after(lines: list[str], idx: int) -> int | None:
    """Line number of the quoted script a narration label introduces, if any."""
    scan = idx + 1
    limit = min(idx + 14, len(lines))
    while scan <= limit:
        text = lines[scan - 1].strip()
        if not text or _STAGE_DIRECTION.match(text):
            scan += 1
            continue
        if _A11Y_NOTE.search(text):
            if text.startswith(">"):
                while scan <= limit and lines[scan - 1].strip().startswith(">"):
                    scan += 1
            else:
                scan += 1
            continue
        if text.startswith("<!--") or text.startswith("-->"):
            scan += 1
            continue
        return scan if text.startswith(">") else None
    return None


# ==========================================================================
# Build / check
# ==========================================================================

def build_outputs(model: dict) -> dict[Path, bytes]:
    recon = reconcile(model)
    model = dict(model)
    model["reconciliation"] = recon
    outputs: dict[Path, bytes] = {
        OUT_JSON: render_manifest_json(model).encode("utf-8"),
        OUT_ASSETS_CSV: render_csv(ASSET_CSV_HEADER, asset_csv_rows(model)),
        OUT_DELIVERABLES_CSV: render_csv(DELIVERABLE_CSV_HEADER, deliverable_csv_rows(model)),
        OUT_REUSE_CSV: render_csv(REUSE_CSV_HEADER, reuse_csv_rows(model)),
        OUT_REGISTER_MD: render_register_md(model).encode("utf-8"),
        OUT_MIGRATION_CSV: render_csv(MIGRATION_CSV_HEADER, recon["rows"]),
        OUT_MIGRATION_MD: render_migration_md(model, recon).encode("utf-8"),
        OUT_XLSX: render_xlsx(model),
    }
    return outputs


def write_outputs(outputs: dict[Path, bytes]) -> list[Path]:
    written = []
    for path, data in outputs.items():
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(data)
        written.append(path)
    return written


def compare_outputs(outputs: dict[Path, bytes]) -> list[str]:
    stale = []
    for path, data in outputs.items():
        if not path.exists():
            stale.append(f"HIÁNYZIK {_rel(path)}")
            continue
        current = path.read_bytes()
        if current != data:
            stale.append(f"ELCSÚSZOTT {_rel(path)} "
                         f"(committed {len(current)} bájt, generált {len(data)} bájt)")
    return stale


# ==========================================================================
# diff against a git ref
# ==========================================================================

def manifest_at_ref(ref: str) -> dict:
    """Compile the manifest as it stands in ``ref`` using a temporary worktree-free checkout."""
    import tempfile
    listing = subprocess.run(
        ["git", "-C", str(ROOT), "ls-tree", "-r", "--name-only", ref, "02 Tervezet"],
        capture_output=True, text=True, check=True).stdout.splitlines()
    with tempfile.TemporaryDirectory() as tmp:
        tmp_root = Path(tmp)
        for rel in listing:
            if not rel.endswith(".md"):
                continue
            blob = subprocess.run(["git", "-C", str(ROOT), "show", f"{ref}:{rel}"],
                                  capture_output=True, check=True).stdout
            target = tmp_root / rel
            target.parent.mkdir(parents=True, exist_ok=True)
            target.write_bytes(blob)
        with _rel_root(tmp_root):
            return compile_manifest(tmp_root / "02 Tervezet", strict=False)


def diff_manifests(old: dict, new: dict) -> list[str]:
    old_by_id = {a["id"]: a for a in old["assets"]}
    new_by_id = {a["id"]: a for a in new["assets"]}
    out = []
    for aid in sorted(set(new_by_id) - set(old_by_id), key=natural_key):
        out.append(f"ADDED         {aid}  ({new_by_id[aid]['title']})")
    for aid in sorted(set(old_by_id) - set(new_by_id), key=natural_key):
        out.append(f"REMOVED       {aid}  ({old_by_id[aid]['title']})")
    for aid in sorted(set(old_by_id) & set(new_by_id), key=natural_key):
        before, after = old_by_id[aid], new_by_id[aid]
        if before["copy_hash"] != after["copy_hash"]:
            out.append(f"COPY_CHANGED  {aid}")
        if before["spec_hash"] != after["spec_hash"]:
            out.append(f"SPEC_CHANGED  {aid}")
        if before["reuse_of"] != after["reuse_of"] or before["mode"] != after["mode"]:
            out.append(f"REUSE_CHANGED {aid}  ({before['mode']}/{before['reuse_of']}"
                       f" → {after['mode']}/{after['reuse_of']})")
    return out


# ==========================================================================
# Selftest
# ==========================================================================

def selftest() -> int:
    failures = []

    def check(name, condition):
        if condition:
            print(f"  ok    {name}")
        else:
            print(f"  FAIL  {name}")
            failures.append(name)

    body = '> „Szia!\n> Gondolj vissza.   \n>\n> És **kész**.”'
    normalised = normalise_source_text(body)
    check("blockquote-prefix eltávolítva", not normalised.startswith(">"))
    check("magyar idézőjel megmarad", "„Szia!" in normalised and "kész**.”" in normalised)
    check("bekezdéshatár megmarad", "\n\n" in normalised)
    check("sorvégi szóköz levágva", "vissza.\n" in normalised)

    lines = ['<!-- @asset', '{"id": "M1.1-VID-01", "kind": "video"}', '-->']
    payload, end = _parse_payload(lines, 0, "asset", "teszt")
    check("többsoros deklaráció", payload["id"] == "M1.1-VID-01" and end == 2)

    lines = ['<!-- @source {"id": "S1", "kind": "narration"} -->']
    payload, end = _parse_payload(lines, 0, "source", "teszt")
    check("egysoros deklaráció", payload["id"] == "S1" and end == 0)

    try:
        _parse_payload(['<!-- @asset', '{oops}', '-->'], 0, "asset", "teszt")
        check("hibás JSON elutasítva", False)
    except ManifestError:
        check("hibás JSON elutasítva", True)

    try:
        _parse_payload(['<!-- @asset', '{"id": "x"}'], 0, "asset", "teszt")
        check("lezáratlan blokk elutasítva", False)
    except ManifestError:
        check("lezáratlan blokk elutasítva", True)

    check("deliverable-ID nem ütközik asset-ID-vel",
          "::" in deliverable_id("M1.1-VID-01", "captions"))
    check("hash determinisztikus",
          sha256_text("alma") == sha256_text("alma") and
          sha256_text("alma") != sha256_text("Alma"))

    unit, module, kind = file_identity(
        ROOT / "02 Tervezet/Modulok/M1/Online leckék/M1.1 – x.md")
    check("lecke-egység", (unit, module, kind) == ("M1.1", "M1", "online-lecke"))
    unit, module, kind = file_identity(ROOT / "02 Tervezet/Modulok/M1/M1 – Vakfolt.md")
    check("hub-egység", (unit, module, kind) == ("M1-HUB", "M1", "hub"))
    unit, module, kind = file_identity(ROOT / "02 Tervezet/Modulok/M1/M1 – Kapu – x.md")
    check("kapu-egység", (unit, module, kind) == ("M1-KAPU", "M1", "kapu"))
    unit, module, kind = file_identity(ROOT / "02 Tervezet/Modulok/M1/Peulák/M1.F – x.md")
    check("peula-egység", (unit, module, kind) == ("M1.F", "M1", "peula"))

    print(f"--- selftest {'OK' if not failures else 'FAILED'} "
          f"({len(failures)} hiba) ---")
    return 1 if failures else 0


# ==========================================================================
# CLI
# ==========================================================================

def _report_errors(errors: list[ManifestError]) -> None:
    for error in errors:
        print(f"HIBA {error}", file=sys.stderr)


def _load(strict: bool = True):
    try:
        return compile_manifest(strict=strict)
    except ManifestValidationFailed as exc:
        _report_errors(exc.errors)
        print(f"\n{len(exc.errors)} manifest-hiba.", file=sys.stderr)
        raise SystemExit(1)


def cmd_validate(_args) -> int:
    model = compile_manifest(strict=False)
    if model["errors"]:
        _report_errors(model["errors"])
        print(f"\n{len(model['errors'])} manifest-hiba.", file=sys.stderr)
        return 1
    stats = compute_stats(model)
    print(f"OK — {stats['assets']} asset, {stats['deliverables']} deliverable, "
          f"{stats['sources']} forrásblokk, {stats['files_discovered']} fájl.")
    return 0


def cmd_build(_args) -> int:
    model = _load()
    outputs = build_outputs(model)
    written = write_outputs(outputs)
    stats = compute_stats(model)
    print(f"Assetek: {stats['assets']} · deliverable-ek: {stats['deliverables']} · "
          f"forrásblokkok: {stats['sources']}")
    for path in written:
        print(f"  írva: {_rel(path)} ({len(outputs[path])} bájt)")
    return 0


def cmd_check(_args) -> int:
    model = _load()
    outputs = build_outputs(model)
    stale = compare_outputs(outputs)
    if stale:
        for entry in stale:
            print(entry, file=sys.stderr)
        print("\nA generált regiszter nem a jelenlegi Markdown-ból származik.",
              file=sys.stderr)
        print("Futtasd: python3 tools/media_manifest.py build", file=sys.stderr)
        return 1
    recon = reconcile(model)
    if recon["unmapped"]:
        print(f"{recon['unmapped']} történeti sor nincs egyeztetve.", file=sys.stderr)
        return 1
    print(f"OK — {len(outputs)} generált kimenet naprakész, "
          f"{recon['legacy_total']} történeti sor egyeztetve.")
    return 0


def cmd_stats(_args) -> int:
    model = _load(strict=False)
    if model["errors"]:
        _report_errors(model["errors"])
    stats = compute_stats(model)
    print(json.dumps(stats, ensure_ascii=False, indent=2))
    return 0


def cmd_reconcile(_args) -> int:
    model = _load()
    recon = reconcile(model)
    print(f"Történeti sor: {recon['legacy_total']} · "
          f"diszpozícionálva: {recon['mapped_total']} · "
          f"nem egyeztetett: {recon['unmapped']}")
    for status in RECON_STATUSES:
        count = recon["tally"].get(status, 0)
        if count:
            print(f"  {status:26s} {count}")
    for conflict in recon["conflicts"]:
        print(f"  ÜTKÖZÉS {conflict}", file=sys.stderr)
    return 1 if (recon["unmapped"] or recon["conflicts"]) else 0


def cmd_lint(args) -> int:
    model = _load(strict=False)
    findings = lint(model)
    high = [f for f in findings if f["confidence"] == "HIGH"]
    for finding in findings:
        if args.high_only and finding["confidence"] != "HIGH":
            continue
        print(f"POSSIBLE_UNDECLARED_ASSET [{finding['confidence']}] "
              f"{finding['file']}:{finding['line']} "
              f"({finding['heading']}) — {finding['reason']} → {finding['fix']}")
    print(f"\n{len(findings)} jelzés, ebből {len(high)} HIGH.")
    return 1 if high else 0


def cmd_diff(args) -> int:
    new = _load(strict=False)
    old = manifest_at_ref(args.ref)
    changes = diff_manifests(old, new)
    for change in changes:
        print(change)
    print(f"\n{len(changes)} változás a(z) {args.ref} óta.")
    return 0


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        prog="media_manifest.py",
        description="Asset Manifest v2 — determinisztikus média-asset fordító.")
    parser.add_argument("--selftest", action="store_true",
                        help="parser/normalizáló füstteszt, korpusz nélkül")
    sub = parser.add_subparsers(dest="command")

    sub.add_parser("build", help="minden generált kimenet újraépítése")
    sub.add_parser("validate", help="séma-, hivatkozás- és a11y-ellenőrzés")
    sub.add_parser("check", help="elcsúszott generált kimenet keresése (CI)")
    sub.add_parser("reconcile", help="a 747 történeti sor egyeztetése")
    sub.add_parser("stats", help="jelenlegi darabszámok JSON-ben")
    lint_parser = sub.add_parser("lint", help="lehetséges deklarálatlan assetek")
    lint_parser.add_argument("--high-only", action="store_true",
                             help="csak a HIGH bizonyosságú jelzések")
    diff_parser = sub.add_parser("diff", help="változások egy git ref óta")
    diff_parser.add_argument("ref")

    args = parser.parse_args(argv)
    if args.selftest:
        return selftest()

    handlers = {
        "build": cmd_build, "validate": cmd_validate, "check": cmd_check,
        "reconcile": cmd_reconcile, "stats": cmd_stats, "lint": cmd_lint,
        "diff": cmd_diff,
    }
    handler = handlers.get(args.command)
    if handler is None:
        parser.print_help()
        return 2
    try:
        return handler(args)
    except ManifestError as exc:
        print(f"HIBA {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
