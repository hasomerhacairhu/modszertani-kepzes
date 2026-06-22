#!/usr/bin/env python3
# Média-asset regiszter → XLSX (többlapos) + CSV-k. Merge a fő + kiegészítő JSON-ból.
import json, csv, sys, re, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = "/Users/heymarcell/DEV/modszertani-kepzes/02 Tervezet/_MÉDIA-ASSETEK"
SRCS = sys.argv[1:]
if not SRCS:
    print("Használat: build-data.py <result1.json> [result2.json ...]"); sys.exit(1)

def load(p):
    d = json.load(open(p, encoding="utf-8"))
    if isinstance(d, dict) and isinstance(d.get("result"), str):
        try: d = json.loads(d["result"])
        except: pass
    if isinstance(d, dict) and isinstance(d.get("result"), dict):
        d = d["result"]
    return d

def clean(s):
    if s is None: return ""
    s = str(s).replace("&amp;", "&")
    s = re.sub(r"\s+", " ", s).strip()
    return s

# ---- merge ----
assets = {}
dedup_groups = []
audit_dims = {}
empty_files = set()
production_rules = []
for p in SRCS:
    d = load(p)
    if d.get("productionRules"):
        production_rules = d["productionRules"]
    for a in d.get("assets", []):
        aid = a.get("assetId")
        if aid and aid not in assets:
            assets[aid] = a
    for grp in d.get("dedup", []):
        bucket = grp.get("bucket", "—")
        for g in grp.get("groups", []):
            dedup_groups.append({"canonicalId": g.get("canonicalId",""), "memberIds": g.get("memberIds",[]),
                                 "bucket": bucket, "reason": g.get("reason",""), "reuseNote": g.get("reuseNote","")})
    for a in d.get("audit", []):
        dim = a.get("dimension","")
        audit_dims.setdefault(dim, {"summary": a.get("summary",""), "findings": []})
        audit_dims[dim]["findings"].extend(a.get("findings", []))
    for mf in d.get("missingFiles", []):
        empty_files.add(clean(mf))

rows = list(assets.values())
files_with = {a.get("file") for a in rows}
empty_files = sorted(f for f in empty_files if f not in files_with and clean(f) not in {clean(x) for x in files_with})

def prov_norm(p):
    p = (p or "").lower()
    if p.startswith("ai") or "ai-gener" in p: return "AI-generált"
    if "stock" in p: return "stock"
    if "vegyes" in p: return "vegyes"
    if any(k in p for k in ["emberi","ember által","narrác","forrásban kész","lektorált"]): return "emberi-felvétel / kész szöveg"
    if "szöveges-ekviv" in p: return "szöveges-ekvivalens (származtatott)"
    return "egyéb/ismeretlen"

KAT = {"digitális-generálandó":"digi","szöveges-ekvivalens":"a11y-szöveg","print-fizikai":"print"}
MODORD = {m:i for i,m in enumerate(["M0","M1","M2","M3","M4","M5","M6","M7","Z"])}
KINDORD = {"online-lecke":0,"peula":1,"hub":2,"kapu":3}

def sortkey(a):
    return (MODORD.get(a.get("mod"),9), KINDORD.get(a.get("kind"),9), str(a.get("fileShort","")),
            [int(t) if t.isdigit() else t for t in re.split(r"(\d+)", str(a.get("assetId","")))])
rows.sort(key=sortkey)

# dedup-ref per asset
dedup_ref = {}
for g in dedup_groups:
    for mid in g["memberIds"]:
        base = re.sub(r"\s*\(.*?\)\s*$", "", str(mid)).strip()
        if base != g["canonicalId"]:
            dedup_ref.setdefault(base, "🔁 = " + g["canonicalId"])

# audit per asset
SEVRANK = {"blokkoló":0,"fontos":1,"javasolt":2}
audit_per = {}
for dim, dd in audit_dims.items():
    for f in dd["findings"]:
        aid = f.get("assetId","")
        audit_per.setdefault(aid, []).append((SEVRANK.get(f.get("severity"),3), f.get("severity",""), f.get("issue",""), f.get("fix",""), dim))

def audit_cell(aid):
    items = sorted(audit_per.get(aid, []))
    ic = {"blokkoló":"⛔","fontos":"⚠️","javasolt":"ℹ️"}
    return " · ".join(f'{ic.get(s,"")} {clean(iss)}' + (f" → {clean(fx)}" if fx else "") for _,s,iss,fx,_ in items)

# ============ XLSX ============
wb = Workbook()
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CAT_FILL = {"digitális-generálandó":PatternFill("solid",fgColor="E2EFDA"),
            "szöveges-ekvivalens":PatternFill("solid",fgColor="FFF2CC"),
            "print-fizikai":PatternFill("solid",fgColor="DEEBF7")}

def style_sheet(ws, headers, widths, wrapcols):
    for c,(h,w) in enumerate(zip(headers,widths),1):
        cell = ws.cell(1,c,h); cell.fill=HDR_FILL; cell.font=HDR_FONT; cell.alignment=Alignment(wrap_text=True,vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.row_dimensions[1].height = 30
    ws._wrapcols = wrapcols

# --- Assetek ---
ws = wb.active; ws.title = "Assetek"
H = ["ID","Modul","Fájltípus","Forrásfájl","Hol (slide/blokk)","Sor / horgony","Asset-típus","Kategória","Legyártandó?",
     "Cím","Mit kell generálni","Miért (cél)","Tech-spec","A11y igény","Eredet (nyers)","Eredet (norm)","Dedup","Audit-jelölés","Megjegyzés","Útvonal"]
W = [16,7,12,34,26,22,18,16,14,24,52,40,28,30,22,20,16,46,28,46]
WRAPC = {5,11,12,13,14,17,18,19}
style_sheet(ws, H, W, WRAPC)
r = 2
for a in rows:
    aid = a.get("assetId","")
    vals = [aid, a.get("mod",""), a.get("kind",""), clean(a.get("fileShort","")), clean(a.get("location","")),
            clean(a.get("lineRef","")), a.get("assetType",""), a.get("category",""),
            ("újrahasznosítás" if dedup_ref.get(aid) else "igen"),
            clean(a.get("title","")), clean(a.get("contentSpec","")), clean(a.get("purpose","")), clean(a.get("techSpec","")),
            clean(a.get("a11y","")), clean(a.get("provenance","")), prov_norm(a.get("provenance","")),
            dedup_ref.get(aid,"") or clean(a.get("dedup","")), audit_cell(aid) or clean(a.get("audit","")), clean(a.get("notes","")), clean(a.get("file",""))]
    for c,v in enumerate(vals,1):
        cell = ws.cell(r,c,v); cell.border=BORDER
        cell.alignment = WRAP if c in WRAPC else TOP
    cat = a.get("category","")
    if cat in CAT_FILL: ws.cell(r,8).fill = CAT_FILL[cat]
    r += 1

# --- Dedup ---
ws2 = wb.create_sheet("Dedup-térkép")
H2 = ["Kanonikus (megtartandó)","Tagok","Bucket","Indok / újrahasznosítás"]; W2=[24,40,22,90]
style_sheet(ws2, H2, W2, {1,2,3,4})
for i,g in enumerate(dedup_groups,2):
    vals=[g["canonicalId"], ", ".join(str(m) for m in g["memberIds"]), g["bucket"],
          clean(g["reason"] + (" — "+g["reuseNote"] if g["reuseNote"] else ""))]
    for c,v in enumerate(vals,1):
        cell=ws2.cell(i,c,v); cell.border=BORDER; cell.alignment=WRAP

# --- Audit ---
ws3 = wb.create_sheet("Audit")
H3=["Súlyosság","Dimenzió","Asset ID","Probléma","Javasolt javítás"]; W3=[12,40,16,70,60]
style_sheet(ws3, H3, W3, {2,4,5})
SEVFILL={"blokkoló":PatternFill("solid",fgColor="F8CBAD"),"fontos":PatternFill("solid",fgColor="FFE699"),"javasolt":PatternFill("solid",fgColor="D9E1F2")}
audit_flat=[]
for dim,dd in audit_dims.items():
    for f in dd["findings"]:
        audit_flat.append((SEVRANK.get(f.get("severity"),3), f.get("severity",""), dim, f.get("assetId",""), f.get("issue",""), f.get("fix","")))
audit_flat.sort()
for i,(_,sev,dim,aid,iss,fx) in enumerate(audit_flat,2):
    vals=[sev,dim,aid,clean(iss),clean(fx)]
    for c,v in enumerate(vals,1):
        cell=ws3.cell(i,c,v); cell.border=BORDER; cell.alignment=WRAP if c in {2,4,5} else TOP
    if sev in SEVFILL: ws3.cell(i,1).fill=SEVFILL[sev]

# --- Médiamentes ---
ws4 = wb.create_sheet("Médiamentes fájlok")
style_sheet(ws4, ["Ellenőrzött, asset nélküli fájl","Indok"], [70,50], {1,2})
for i,f in enumerate(empty_files,2):
    ws4.cell(i,1,f).alignment=WRAP
    ws4.cell(i,2,"Item-bank / áttekintő / reflexiós fájl — a videó/kép-utalások tanulói beadványra vagy más fájl már katalogizált assetjére mutatnak").alignment=WRAP

# --- Produkciós konvenciók ---
if production_rules:
    ws6 = wb.create_sheet("Produkciós konvenciók")
    style_sheet(ws6, ["Szabály","Megnevezés","Tartalom"], [10,30,120], {2,3})
    for i,rl in enumerate(production_rules,2):
        c=ws6.cell(i,1,rl["id"]); c.font=Font(bold=True); c.alignment=TOP; c.border=BORDER
        ws6.cell(i,2,rl["title"]).alignment=WRAP; ws6.cell(i,2).border=BORDER
        ws6.cell(i,3,rl["text"]).alignment=WRAP; ws6.cell(i,3).border=BORDER

# --- Összesítő ---
ws5 = wb.create_sheet("Összesítő"); ws5.sheet_view.showGridLines=False
def cnt(key):
    d={}
    for a in rows: d[a.get(key,"")]=d.get(a.get(key,""),0)+1
    return d
dupc = sum(max(0,len(set(g["memberIds"]))-1) for g in dedup_groups)
title_f=Font(bold=True,size=13); sub_f=Font(bold=True,size=11,color="1F4E78")
ws5["A1"]="Média-asset regiszter — összesítő"; ws5["A1"].font=title_f
def block(start,title,d,order=None):
    ws5.cell(start,1,title).font=sub_f
    items = [(k,d[k]) for k in order if k in d] if order else sorted(d.items(),key=lambda x:-x[1])
    rr=start+1
    for k,v in items:
        ws5.cell(rr,1,k); ws5.cell(rr,2,v); rr+=1
    return rr+1
row=3
ws5.cell(row,1,"Összes asset-sor"); ws5.cell(row,2,len(rows)); row+=1
ws5.cell(row,1,"Egyedi, ténylegesen legyártandó"); ws5.cell(row,2,len(rows)-dupc); row+=1
ws5.cell(row,1,"Dedup-újrahasznosítás"); ws5.cell(row,2,dupc); row+=1
ws5.cell(row,1,"Lefedett forrásfájl"); ws5.cell(row,2,len(files_with)); row+=1
ws5.cell(row,1,"Médiamentes (ellenőrzött) fájl"); ws5.cell(row,2,len(empty_files)); row+=2
row=block(row,"Kategória szerint",cnt("category"))
row=block(row,"Asset-típus szerint",cnt("assetType"))
row=block(row,"Modul szerint",cnt("mod"),order=["M0","M1","M2","M3","M4","M5","M6","M7","Z"])
prov={}
for a in rows: prov[prov_norm(a.get("provenance",""))]=prov.get(prov_norm(a.get("provenance","")),0)+1
row=block(row,"Eredet (gyártási mód) szerint",prov)
sev={}
for _,s,_,_,_,_ in audit_flat: sev[s]=sev.get(s,0)+1
row=block(row,"Audit súlyosság szerint",sev,order=["blokkoló","fontos","javasolt"])
ws5.column_dimensions["A"].width=40; ws5.column_dimensions["B"].width=12

# sheet-sorrend: Összesítő elöl
wb.move_sheet("Összesítő", -(len(wb.sheetnames)-1))

xlsx_path = os.path.join(ROOT, "Média-asset regiszter.xlsx")
wb.save(xlsx_path)

# ============ CSV-k (UTF-8 BOM, hogy Excel HU is jól nyissa) ============
def write_csv(name, header, data):
    path = os.path.join(ROOT, name)
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(header); w.writerows(data)
    return path

assets_csv = write_csv("assetek.csv", H,
    [[a.get("assetId",""),a.get("mod",""),a.get("kind",""),clean(a.get("fileShort","")),clean(a.get("location","")),
      clean(a.get("lineRef","")),a.get("assetType",""),a.get("category",""),
      ("újrahasznosítás" if dedup_ref.get(a.get("assetId","")) else "igen"),
      clean(a.get("title","")),clean(a.get("contentSpec","")),clean(a.get("purpose","")),clean(a.get("techSpec","")),
      clean(a.get("a11y","")),clean(a.get("provenance","")),prov_norm(a.get("provenance","")),
      dedup_ref.get(a.get("assetId",""),"") or clean(a.get("dedup","")),audit_cell(a.get("assetId","")) or clean(a.get("audit","")),clean(a.get("notes","")),clean(a.get("file",""))] for a in rows])
dedup_csv = write_csv("dedup.csv", H2,
    [[g["canonicalId"], ", ".join(str(m) for m in g["memberIds"]), g["bucket"], clean(g["reason"]+(" — "+g["reuseNote"] if g["reuseNote"] else ""))] for g in dedup_groups])
audit_csv = write_csv("audit.csv", H3,
    [[s,dim,aid,clean(iss),clean(fx)] for _,s,dim,aid,iss,fx in audit_flat])

print(f"XLSX: {xlsx_path}")
print(f"  munkalapok: {wb.sheetnames}")
print(f"  assetek: {len(rows)} | dedup: {len(dedup_groups)} | audit: {len(audit_flat)} | médiamentes: {len(empty_files)}")
print(f"CSV: {os.path.basename(assets_csv)}, {os.path.basename(dedup_csv)}, {os.path.basename(audit_csv)}")
